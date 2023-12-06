from openpyxl import load_workbook

root: str = "data/PRICE-LIST TS-2022.xlsx"
work_book: str = "data/DB_bending.xlsx"


wb_1 = load_workbook(root)

wb_2 = load_workbook(work_book)

sheet_1 = wb_1["LISTINO 2022"]

sheet_2 = wb_2["Матриця багаторучова"]
wb_2.active
weight = None
price = None

for index in range(2, sheet_1.max_row + 1):
    for i in range(2, sheet_2.max_row + 1):
        item = sheet_2["C" +str(i)].value
        item = item.replace("K", "F")
        if sheet_1["B" +str(index)].value == item:
            sheet_2["H" + str(i)] = sheet_1["G" +str(index)].value
            sheet_2["I" + str(i)] = sheet_1["E" +str(index)].value


wb_2.save("data/DB_bending_1.xlsx")