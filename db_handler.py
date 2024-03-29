import openpyxl
from openpyxl import *
from openpyxl.utils import get_column_letter

DB_PATH = "data/DB_bending.xlsx"
COMMERCIAL_OFFER_EMPTY_SAMPLE_PATH = ""

CALCULATION_EMPTY_SMPLE_PATH = ""


class MyDb:

    def __init__(self):
        self.path_db = DB_PATH

    def get_type_holder_list(self) -> list:
        pass

    def get_type_item_list(self) -> list:
        pass

    @staticmethod
    def open_book(path_book: str = DB_PATH):
        """
        Метод повертає об'єкт Workbook, який був створений при
        відкритті файла типу Excel за розташуванням парамерта
        path_book
        :param path_book: str
        :return: Workbook
        """
        book = load_workbook(path_book)
        return book


    @staticmethod
    def get_code_list(holder_item: tuple) -> tuple:
        holder: str = holder_item[0]
        item: str = holder_item[1]
        wb = load_workbook(DB_PATH)
        code_list: list = [" "]
        work_sheet = wb[item]
        max_row_item = work_sheet.max_row
        for i in range(1, max_row_item + 1):
            if work_sheet["B"+str(i)].value == holder:
                if len(work_sheet["C"+str(i)].value) == 6:
                    code_list.append(work_sheet["C"+str(i)].value)
                if len(work_sheet["C"+str(i)].value) == 7:
                    code_list.append(work_sheet["C"+str(i)].value[0:6])
                if len(work_sheet["C"+str(i)].value) == 8:
                    code = (work_sheet["C" + str(i)].value[0:6] +
                            work_sheet["C" + str(i)].value[-1])
                    code_list.append(code)
        result_list = list(set(code_list))
        result_list.sort()
        del code_list
        return tuple(result_list)

    @staticmethod
    def get_length_item(holder_item_code: tuple) -> tuple:
        wb = load_workbook(DB_PATH)
        length_list: list = [" "]
        holder: str = holder_item_code[0]
        item: str = holder_item_code[1]
        code: str = holder_item_code[2]
        work_sheet = wb[item]
        max_row_item = work_sheet.max_row
        if holder_item_code[2][-1] == "X":
            for i in range(1, max_row_item + 1):
                if (work_sheet["B" + str(i)].value == holder and
                        work_sheet["C" + str(i)].value[0:6] == code[0:6] and
                        work_sheet["C" + str(i)].value[-1] == "X"
                        ):
                    length_list.append(work_sheet["G" + str(i)].value)
            # return length_list
        else:
            for i in range(1, max_row_item + 1):
                if (work_sheet["B"+str(i)].value == holder and
                        work_sheet["C"+str(i)].value[0:6] == code and
                        work_sheet["C" + str(i)].value[-1] != "X"
                ):
                    length_list.append(work_sheet["G"+str(i)].value)
            # return length_list
        return tuple(length_list)
    @staticmethod
    def get_code_length(
                        item: str,
                        short_code: str,
                        length: str) -> str:
        """Метод повертає повний код виробу якщо остання літера назви
         X та початок збігається з short_code """
        wb = load_workbook(DB_PATH)
        work_sheet = wb[item]
        max_row_item = work_sheet.max_row
        result: str = "Empty"
        for i in range(1, max_row_item + 1):

            if (work_sheet["C" + str(i)].value[-1] == "X" and
                    work_sheet["C" + str(i)].value[0:6] == short_code[0:6] and
                    str(work_sheet["G" + str(i)].value) == length):

                result = work_sheet["C" + str(i)].value
                return result
        return result

    @staticmethod
    def get_en_description(data_list: list) -> str:
        en_description = "en_description"
        holder: str = data_list[0]
        item: str = data_list[1]
        code: str = data_list[2]
        length: str = data_list[3]
        full_code: str = data_list[4]
        wb = load_workbook(DB_PATH)
        work_sheet = wb[item]
        max_row_item = 0
        max_row_item = work_sheet.max_row
        for i in range(1, max_row_item + 1):
            if work_sheet["C"+str(i)].value == full_code:
                return work_sheet["D"+str(i)].value
        return en_description

    @staticmethod
    def get_ua_description(data_list: list) -> str:
        ua_description = "ua_description"
        holder: str = data_list[0]
        item: str = data_list[1]
        code: str = data_list[2]
        length: str = data_list[3]
        full_code: str = data_list[4]
        wb = load_workbook(DB_PATH)
        work_sheet = wb[item]
        max_row_item = 0
        max_row_item = work_sheet.max_row
        for i in range(1, max_row_item + 1):
            if work_sheet["C"+str(i)].value == full_code:
                return work_sheet["E"+str(i)].value
        return ua_description
    @staticmethod
    def get_length(length: str) -> str:
        # if length in ["+", "="]:
        #     pass
        if "=" in str(length):
            return length[-3:]

        return length


    def get_item(self, parameters_list: list) -> list:
        type_loder: str = parameters_list[0]
        item: str = parameters_list[1]
        code: str = parameters_list[2]
        length_item: str = MyDb.get_length(parameters_list[3])

    @staticmethod
    def get_full_code_item(parameters: list) -> str:
        full_code: str = ""
        holder: str = parameters[0]
        item: str = parameters[1]
        code: str = parameters[2]
        length: str = parameters[3]
        wb = load_workbook(DB_PATH)
        work_sheet = wb[item]
        max_row_item = 0
        max_row_item = work_sheet.max_row

        if code[-1] == "X":
            full_code = MyDb.get_code_length(item, code, length)
        else:
            for i in range(1, max_row_item + 1):
                if work_sheet["C"+str(i)].value[0:6] == code and \
                        str(work_sheet["G" + str(i)].value) == length:

                    return work_sheet["C" + str(i)].value
        return full_code

    @staticmethod
    def get_info_item(data_list: list) -> dict:
        info_item: dict = {}
        wb = load_workbook(DB_PATH)
        work_sheet = wb[data_list[1]]
        max_row_item = work_sheet.max_row
        max_column_item = work_sheet.max_column
        for i in range(1, max_row_item + 1):
            if work_sheet["C" + str(i)].value == data_list[-1]:
                info_item["type_holder"] = work_sheet["B" + str(i)].value
                info_item["item"] = work_sheet.title
                info_item["code_item"] = work_sheet["C" + str(i)].value
                info_item["en_name_item"] = work_sheet["D" + str(i)].value
                info_item["ua_name_item"] = work_sheet["E" + str(i)].value
                info_item["image_path"] = work_sheet["F" + str(i)].value
                info_item["length_item"] = str(work_sheet["G" + str(i)].value)
                info_item["weight"] = work_sheet["H" + str(i)].value
                info_item["price_item"] = work_sheet["I" + str(i)].value
                info_item["parameters"] = {}
                for j in range(9, max_column_item):
                    info_item["parameters"][work_sheet[chr(65 + j) + "1"].value] = work_sheet[chr(65 + j) + str(i)].value

        return info_item

    @staticmethod
    def get_punch_by_holder(book=None, holder=None) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу
        :param book: Workbook
        :param holder: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        for index in range(1, max_row_item_punch):
            if work_sheet_punch["B" + str(index)].value == holder:
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                            work_sheet_punch["C" + str(index)].value[0:6] +
                            work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_angle(
            book=None,
            type_holder=None,
            angle=None) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу та
         певного кута
        :param book: Workbook
        :param type_holder: str
        :param angle: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")

        for index in range(1, max_row_item_punch):

            if (
                    work_sheet_punch["B" + str(index)].value ==
                    type_holder and
                    str(work_sheet_punch["J" + str(index)].value) == angle
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_height(book=None,
                                   type_holder=None,
                                   height=None) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу та
         певної висоти
        :param book: Workbook
        :param type_holder: str
        :param height: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        height = height.replace(",", ".")
        for index in range(1, max_row_item_punch):
            if (
                    work_sheet_punch["B" + str(index)].value == type_holder
                    and str(work_sheet_punch["K" + str(index)].value) == height
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_radius(book=None,
                                   type_holder=None,
                                   radius=None) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу та
         певної висоти
        :param book: Workbook
        :param type_holder: str
        :param radius: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        radius = radius.replace(",",  ".")
        for index in range(1, max_row_item_punch):
            if (
                    work_sheet_punch["B" + str(index)].value == type_holder
                    and str(work_sheet_punch["L" + str(index)].value) == radius
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_angle_height(
        book=None,
        type_holder=None,
        angle=None,
        height=None
    ) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу,
         певного кута та певної висоти
        :param book: Workbook
        :param type_holder: str
        :param angle: str
        :param height: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        height = height.replace(",", ".")
        for index in range(1, max_row_item_punch):

            if (
                    work_sheet_punch["B" + str(index)].value ==
                    type_holder and
                    str(work_sheet_punch["J" + str(index)].value) == angle
                    and str(work_sheet_punch["K" + str(index)].value) == height
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_angle_radius(
        book=None,
        type_holder=None,
        angle=None,
        radius=None
    ) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу,
         певного кута та певного радіуса
        :param book: Workbook
        :param type_holder: str
        :param angle: str
        :param radius: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        radius = radius.replace(",", ".")
        for index in range(1, max_row_item_punch):

            if (
                    work_sheet_punch["B" + str(index)].value ==
                    type_holder and
                    str(work_sheet_punch["J" + str(index)].value) == angle
                    and str(work_sheet_punch["L" + str(index)].value) == radius
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_punch_by_holder_height_radius(
        book=None,
        type_holder=None,
        height=None,
        radius=None
    ) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу,
         певної висоти та певного радіуса
        :param book: Workbook
        :param type_holder: str
        :param height: str
        :param radius: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        radius = radius.replace(",", ".")
        height = height.replace(",", ".")
        for index in range(1, max_row_item_punch):
            if (
                    work_sheet_punch["B" + str(index)].value ==
                    type_holder and
                    str(work_sheet_punch["K" + str(index)].value) == height
                    and str(work_sheet_punch["L" + str(index)].value) == radius
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)

    def get_punch_by_holder_angle_height_radius(
        book=None,
        type_holder=None,
        angle=None,
        height=None,
        radius=None
    ) -> tuple:
        """
        Функція вертає кортеж кодів усіх пуансонів певного типу,
         певного кута, певної висоти та певного радіуса
        :param book: Workbook
        :param type_holder: str
        :param angle: str
        :param height: str
        :param radius: str
        :return: tuple
        """
        work_sheet_punch = book["Пуансон"]
        max_row_item_punch = work_sheet_punch.max_row
        result_set = set()
        result_set.add(" ")
        radius = radius.replace(",", ".")
        height = height.replace(",", ".")
        for index in range(1, max_row_item_punch):
            if (
                    work_sheet_punch["B" + str(index)].value ==
                    type_holder and
                    str(work_sheet_punch["J" + str(index)].value) == angle
                    and str(work_sheet_punch["K" + str(index)].value) == height
                    and str(work_sheet_punch["L" + str(index)].value) == radius
            ):
                len_code = len(work_sheet_punch["C" + str(index)].value)
                if len_code == 7:
                    result_set.add(
                        str(work_sheet_punch["C" + str(index)].value[0:6])
                    )
                elif len_code == 8:
                    code: str = str(
                        work_sheet_punch["C" + str(index)].value[0:6] +
                        work_sheet_punch["C" + str(index)].value[-1])
                    result_set.add(code)
        result_set = sorted(result_set)
        return tuple(result_set)


    @staticmethod
    def get_punch_code_image(book, code) -> str:
        """
        Фунція вертає назву зображення пуансона згідно кода
        :param book:
        :param code:
        :return:
        """
        sheet = book["Пуансон"]
        max_row_item_punch = sheet.max_row
        if len(code) == 6:
            for index in range(2, max_row_item_punch):
                if sheet["C" + str(index)].value[0:6] == code:
                    return str(sheet["F" + str(index)].value)
        elif len(code) == 7:
            for index in range(2, max_row_item_punch):
                if (
                        sheet["C" + str(index)].value[0:6] == code[0:6]
                        and sheet["C" + str(index)].value[-1] == code[-1]
                ):
                    return str(sheet["F" + str(index)].value)


    @staticmethod
    def get_punch_info(sheet, code_item: str) -> str:
        """
        Функція вертає кортеж
        (
            кут,
            висота,
            радіус,
            Т/Mt
        )
        :param code_item: str
        :return: tuple
        """
        sheet_max_row = sheet.max_row
        print(code_item)
        for index in range(2, sheet_max_row):
            if sheet["C" + str(index)].value[0:6] == code_item or (len(code_item) == 7 and sheet["C" + str(index)].value[0:6] == code_item[0:6]):
                result = ""
                result += chr(int("03B1", 16))
                result += " = "
                result += str(sheet["J" + str(index)].value)
                result += u"\u00b0"
                result += f', H = {str(sheet["K" + str(index)].value)} мм'
                result += f', R = {str(sheet["L" + str(index)].value)} мм'
                result += f', {str(sheet["M" + str(index)].value)} T/м.'

                return result


        return "0, 0, 0, 0"

    @staticmethod
    def get_length_tuple(sheet, code_item: str) -> tuple:
        """
        Функція вертає кортеж довжин певного кода
        пуансону
        :param sheet:
        :param code_item: str
        :return:
        """
        result = []
        max_index = sheet.max_row
        if len(code_item) == 6:
            for index in range(2, max_index):
                if sheet["C" + str(index)].value[0: 6] == code_item:
                    punch_length = str(sheet["G" + str(index)].value)
                    if "=" in punch_length:
                        number_sectioned = punch_length.split("=")
                        result.append(str(number_sectioned[1]).strip() + " SEC")
                    else:
                        result.append(str(sheet["G" + str(index)].value))
                if sheet["C" + str(index - 1)].value[0: 6] == code_item and sheet["C" + str(index)].value[0: 6] != code_item:
                    break
        elif len(code_item) == 7:
            for index in range(2, max_index):
                if sheet["C" + str(index)].value[0: 6] == code_item[0: 6] \
                        and sheet["C" + str(index)].value[-1] == code_item[-1]:
                    punch_length = str(sheet["G" + str(index)].value)
                    if "=" in punch_length:
                        number_sectioned = punch_length.split("=")
                        result.append(str(number_sectioned[1]).strip() + " SEC")
                    else:
                        result.append(str(sheet["G" + str(index)].value))
                if sheet["C" + str(index - 1)].value[0: 6] == code_item and \
                        sheet["C" + str(index-1)].value[-1] == code_item[-1] and \
                        sheet["C" + str(index)].value[0: 6] != code_item and \
                        sheet["C" + str(index)].value[-1] != code_item[-1]:
                    break
        return tuple(result)

    @staticmethod
    def get_die_by_holder(book=None, holder=None ) -> tuple:
        """
        Функція повертає кортеж номрів усіх матриць які
        належать до одного типу тримача
        :return:
        """
        work_sheet_die = book["Матриця одноручова"]
        max_row_item_punch = work_sheet_die.max_row
        result_set = set()
        result_set.add(" ")

        for index in range(2, max_row_item_punch):
            if work_sheet_die["B" + str(index)].value == holder:
                result_set.add(
                    str(work_sheet_die["C" + str(index)].value[0:6])
                )

        if holder == "Amada-promecam":
            work_sheet_die = book["Матриця багаторучова"]
            rows = work_sheet_die.max_row
            for index in range(2, rows + 1):
                if work_sheet_die["B" + str(index)].value == holder:
                    result_set.add(
                        str(work_sheet_die["C" + str(index)].value[0:6])
                    )
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_code_image(book, code) -> str:
        """
        Фунція вертає назву зображення матриці згідно кода
        :param book:
        :param code: str код матриці
        :return:
        """
        sheet_die = book["Матриця одноручова"]
        max_row_die = sheet_die.max_row

        for index in range(2, max_row_die + 1):
            if sheet_die["C" + str(index)].value[0:6] == code:
                return str(sheet_die["F" + str(index)].value)

        sheet_die = book["Матриця багаторучова"]
        max_row_item_die = sheet_die.max_row
        for index in range(2, max_row_item_die + 1):
            if sheet_die["C" + str(index)].value[0:6] == code:
                return str(sheet_die["F" + str(index)].value)

    @staticmethod
    def get_length_die_tuple(book, code_die, holder_die) -> tuple:
        """
        Функція повертае  перелік усіх довжин ватриці певного кода
        :param book: Workbook
        :param code:  str
        :return:
        """
        work_sheet_die = book["Матриця одноручова"]
        max_row_item_die = work_sheet_die.max_row
        result = []

        for index in range(2, max_row_item_die):
            if (
                    work_sheet_die["C" + str(index - 1)].value[0:6] == code_die
                    and work_sheet_die["C" + str(index)].value[0:6] != code_die
            ):
                break
            if work_sheet_die["C" + str(index)].value[0:6] == code_die:
                die_length = str(work_sheet_die["G" + str(index)].value)
                if "=" in die_length:
                    number_sectioned = die_length.split("=")
                    result.append(str(number_sectioned[1]).strip() + " SEC")
                else:
                    result.append(str(work_sheet_die["G" + str(index)].value))
        if len(result) == 0 and holder_die == "Amada-promecam":
            work_sheet_die = book["Матриця багаторучова"]
            rows = work_sheet_die.max_row
            for index in range(2, rows + 1):
                if (
                        work_sheet_die["C" + str(index - 1)].value[0:6] == code_die
                        and work_sheet_die["C" + str(index)].value[0:6] != code_die
                ):
                    break
                if work_sheet_die["C" + str(index)].value[0:6] == code_die:
                    die_length = str(work_sheet_die["G" + str(index)].value)
                    if "=" in die_length:
                        number_sectioned = die_length.split("=")
                        result.append(str(number_sectioned[1]).strip() + " SEC")
                    else:
                        result.append(str(work_sheet_die["G" + str(index)].value))

        return tuple(result)

    def get_die_info(book, code_die) -> str:
        """
        Функція вертає строку по коду code_ide
        (
            кут,
            висота,
            радіус,
            Т/Mt
        )
        :param code_item: str
        :return: tuple
        """
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result: str = ""
        for index in range(2, die_max_row):
            if sheet_die["C" + str(index)].value[0:6] == code_die:
                result += chr(int("03B1", 16))
                result += " = "
                result += str(sheet_die["J" + str(index)].value)
                result += u"\u00b0"
                result += f', V = {str(sheet_die["K" + str(index)].value)} мм'
                result += f', R = {str(sheet_die["M" + str(index)].value)} мм,\n'
                result += f'H = {str(sheet_die["L" + str(index)].value)} мм'
                result += f', {str(sheet_die["N" + str(index)].value)} T/м.'
                return result

        if result != "":
            return result

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row
        result: str = ""

        for index in range(2, die_max_row):
            if sheet_die["C" + str(index)].value[0:6] == code_die:
                if (str(sheet_die["T" + str(index)].value) != "0"
                        and str(sheet_die["S" + str(index)].value) != "0"):
                    result += chr(int("03B1", 16))
                    result += "1"
                    result += " = "
                    result += f'{str(sheet_die["J" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f', V1 = {str(sheet_die["K" + str(index)].value)} мм,'
                    result += f' R1 = {str(sheet_die["L" + str(index)].value)} мм,\n'
                    result += chr(int("03B1", 16))
                    result += "2"
                    result += " = "
                    result += f'{str(sheet_die["M" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f', V2 = {str(sheet_die["N" + str(index)].value)} мм,'
                    result += f' R2 = {str(sheet_die["O" + str(index)].value)} мм,'
                    result += "\n"
                    result += chr(int("03B1", 16))
                    result += "3"
                    result += " = "
                    result += f'{str(sheet_die["P" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f' V3 = {str(sheet_die["Q" + str(index)].value)} мм,'
                    result += f' R3 = {str(sheet_die["R" + str(index)].value)} мм,\n'
                    result += chr(int("03B1", 16))
                    result += "4"
                    result += " = "
                    result += f'{str(sheet_die["S" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f' V4 = {str(sheet_die["T" + str(index)].value)} мм,'
                    result += f' R4 = {str(sheet_die["U" + str(index)].value)} мм,\n'
                    result += f' H =  {str(sheet_die["V" + str(index)].value)} мм,'
                    result += f' {str(sheet_die["W" + str(index)].value)} T/м.'
                    return result

                if (str(sheet_die["J" + str(index)].value) != "0"
                        and str(sheet_die["M" + str(index)].value) != "0"
                        and str(sheet_die["P" + str(index)].value) != "0"
                        and str(sheet_die["T" + str(index)].value) == "0"
                        and str(sheet_die["S" + str(index)].value) == "0"):
                    result += chr(int("03B1", 16))
                    result += "1"
                    result += " = "
                    result += f'{str(sheet_die["J" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f', V1 = {str(sheet_die["K" + str(index)].value)} мм,'
                    result += f' R1 = {str(sheet_die["L" + str(index)].value)} мм,\n'
                    result += chr(int("03B1", 16))
                    result += "2"
                    result += " = "
                    result += f'{str(sheet_die["M" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f', V2 = {str(sheet_die["N" + str(index)].value)} мм,'
                    result += f' R2 = {str(sheet_die["O" + str(index)].value)} мм,\n'
                    result += chr(int("03B1", 16))
                    result += "3"
                    result += " = "
                    result += f'{str(sheet_die["P" + str(index)].value)}'
                    result += u"\u00b0"
                    result += f' V3 = {str(sheet_die["Q" + str(index)].value)} мм,'
                    result += f' R3 = {str(sheet_die["R" + str(index)].value)} мм,\n'
                    result += f' H =  {str(sheet_die["V" + str(index)].value)} мм,'
                    result += f' {str(sheet_die["W" + str(index)].value)} T/м.'
                    return result

                if (str(sheet_die["J" + str(index)].value) == "0"
                        and str(sheet_die["M" + str(index)].value) == "0"
                        and str(sheet_die["P" + str(index)].value) == "0"
                        and str(sheet_die["S" + str(index)].value) == "0"
                        and str(sheet_die["T" + str(index)].value) == "0"
                ):
                    result += f'V1 = {str(sheet_die["K" + str(index)].value)} мм,'
                    result += f' R1 = {str(sheet_die["L" + str(index)].value)} мм,'
                    result += f' V2 = {str(sheet_die["N" + str(index)].value)} мм,\n'
                    result += f' R2 = {str(sheet_die["O" + str(index)].value)} мм,'
                    result += f' V3 = {str(sheet_die["Q" + str(index)].value)} мм,'
                    result += f' R3 = {str(sheet_die["R" + str(index)].value)} мм,\n'
                    result += f' H =  {str(sheet_die["V" + str(index)].value)} мм,'
                    result += f' {str(sheet_die["W" + str(index)].value)} T/м.'
                    return result

                if (str(sheet_die["J" + str(index)].value) != "0"
                        and str(sheet_die["M" + str(index)].value) != "0"
                        and str(sheet_die["P" + str(index)].value) == "0"
                        and str(sheet_die["Q" + str(index)].value) == "0"
                        and str(sheet_die["T" + str(index)].value) == "0"
                        and str(sheet_die["S" + str(index)].value) == "0"):
                    result += chr(int("03B1", 16))
                    result += "1"
                    result += " = "
                    result += str(sheet_die["J" + str(index)].value)
                    result += u"\u00b0"
                    result += f', V1 = {str(sheet_die["K" + str(index)].value)} мм'
                    result += f', R1 = {str(sheet_die["L" + str(index)].value)} мм,\n'
                    result += chr(int("03B1", 16))
                    result += "2"
                    result += " = "
                    result += str(sheet_die["M" + str(index)].value)
                    result += u"\u00b0"
                    result += f', V2 = {str(sheet_die["N" + str(index)].value)} мм'
                    result += f', R2 = {str(sheet_die["O" + str(index)].value)} мм,\n'
                    result += f'H =  {str(sheet_die["V" + str(index)].value)} мм,'
                    result += f' {str(sheet_die["W" + str(index)].value)} T/м.'
                    return result

        return "Помилка у  db_handler.get_die_info"

    @staticmethod
    def get_die_by_holder_angle(book, type_holder, angle) -> tuple:
        """
        Функція повертає кортеж матрць з певним тримачем
        та певним кутом
        :param book:
        :param type_holder:
        :param angle:
        :return:
        """
        angle = int(angle)
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (sheet_die["B"+str(index)].value == type_holder
                    and sheet_die["J"+str(index)].value == angle):
                result_set.add(sheet_die["C"+str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and (
                    sheet_die["J"+str(index)].value == angle
                    or sheet_die["M"+str(index)].value == angle
                    or sheet_die["P"+str(index)].value == angle
                    or sheet_die["S"+str(index)].value == angle
                )
            ):
                result_set.add(sheet_die["C"+str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    def get_die_by_holder_height(
                    book,
                    type_holder,
                    height
                ) -> tuple:
        """
        Функція повертає кортеж матрць з певним тримачем
        та певною висотою
        :param book:
        :param type_holder:
        :param height:
        :return:
        """

        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (sheet_die["B"+str(index)].value == type_holder
                    and str(sheet_die["L"+str(index)].value) == height):
                result_set.add(sheet_die["C"+str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and str(sheet_die["L"+str(index)].value) == height
            ):
                result_set.add(sheet_die["C"+str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_by_holder_distance(
            book,
            type_holder,
            distance
    ) -> tuple:
        """
        Фнкція вертає кортеж з кодів матриць які обрані
        за тримачем та розкриттям
        :param book:
        :param type_holder:
        :param distance:
        :return:
        """
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and str(sheet_die["K"+str(index)].value) == distance
            ):
                result_set.add(sheet_die["C"+str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and (
                    str(sheet_die["K"+str(index)].value) == distance
                    or str(sheet_die["N"+str(index)].value) == distance
                    or str(sheet_die["Q"+str(index)].value) == distance
                    or str(sheet_die["T"+str(index)].value) == distance
                )
            ):
                result_set.add(sheet_die["C"+str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_by_holder_angle_height(book, type_holder, angle, height) -> tuple:
        """
        Функція вертає кортеж матриць певного тимача, кута та висоти
        :return: tuple
        """
        angle = int(angle)
        height = int(height)
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (sheet_die["B"+str(index)].value == type_holder
                    and sheet_die["J"+str(index)].value == angle
                    and sheet_die["L"+str(index)].value == height):
                result_set.add(sheet_die["C"+str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (sheet_die["V"+str(index)].value == height
                    and (sheet_die["J"+str(index)].value == angle
                         or sheet_die["M"+str(index)].value == angle
                         or sheet_die["P"+str(index)].value == angle
                         or sheet_die["T"+str(index)].value == angle)):
                result_set.add(sheet_die["C"+str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_by_holder_angle_distance(
            book,
            type_holder,
            angle,
            distance
    ) -> tuple:
        """
        Функція вертає кортеж матриць певного тимача, кута та розкриття
        """

        angle = int(angle)
        distance = int(distance)

        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (sheet_die["B"+str(index)].value == type_holder
                    and sheet_die["J"+str(index)].value == angle
                    and sheet_die["K"+str(index)].value == distance):
                result_set.add(sheet_die["C"+str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    (sheet_die["K"+str(index)].value == distance
                     and sheet_die["J"+str(index)].value == angle)
                    or (sheet_die["N"+str(index)].value == distance
                        and sheet_die["M"+str(index)].value == angle)
                    or (sheet_die["Q"+str(index)].value == distance
                        and sheet_die["P"+str(index)].value == angle)
                    or (sheet_die["T"+str(index)].value ==  distance
                        and sheet_die["S"+str(index)].value == angle)):
                result_set.add(sheet_die["C"+str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_by_holder_height_distance(
            book,
            type_holder,
            height,
            distance) -> tuple:
        """
        Функція вертає кортеж матриць певного тимача, висота та розкриття
        """
        distance = int(distance)
        height = int(height)
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")

        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and sheet_die["L"+str(index)].value == height
                    and sheet_die["K"+str(index)].value == distance
            ):
                result_set.add(sheet_die["C" + str(index)].value[0:6])

        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    sheet_die["V"+str(index)].value == height and
                    (sheet_die["K"+str(index)].value == distance
                     or sheet_die["N"+str(index)].value == distance
                     or sheet_die["Q"+str(index)].value == distance
                     or sheet_die["T"+str(index)].value == distance)
            ):
                result_set.add(sheet_die["C" + str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)

    @staticmethod
    def get_die_by_holder_ang_hei_dist(
        book,
        type_holder,
        angle,
        height,
        distance) -> tuple:
        """
        Функція вертає кортеж номерів матриць за тримачем, кутом,
        висотою та розкриттям
        :param book:
        :param type_holder:
        :param angle:
        :param height:
        :param distance:
        :return:
        """
        angle = int(angle)
        distance = int(distance)
        height = int(height)
        sheet_die = book["Матриця одноручова"]
        die_max_row = sheet_die.max_row
        result_set: set = set()
        result_set.add("")
        for index in range(2, die_max_row):
            if (
                    sheet_die["B"+str(index)].value == type_holder
                    and sheet_die["J"+str(index)].value == angle
                    and sheet_die["L"+str(index)].value == height
                    and sheet_die["K"+str(index)].value == distance
            ):
                result_set.add(sheet_die["C" + str(index)].value[0:6])
        if type_holder != "Amada-promecam":
            result_set = sorted(result_set)
            return tuple(result_set)

        sheet_die = book["Матриця багаторучова"]
        die_max_row = sheet_die.max_row

        for index in range(2, die_max_row):
            if (
                    sheet_die["V"+str(index)].value == height and
                    ((sheet_die["K"+str(index)].value == distance and
                      sheet_die["J"+str(index)].value == angle)
                     or (sheet_die["N"+str(index)].value == distance and
                         sheet_die["M"+str(index)].value == angle)
                     or (sheet_die["Q"+str(index)].value == distance and
                         sheet_die["P"+str(index)].value == angle)
                     or (sheet_die["T"+str(index)].value == distance and
                         sheet_die["S"+str(index)].value == angle))
            ):
                result_set.add(sheet_die["C" + str(index)].value[0:6])
        result_set = sorted(result_set)
        return tuple(result_set)
    @staticmethod
    def get_all_die_parameters(book, type_holder) -> tuple:
        """
        Функція повертає усі можливі кути, висоти та розкриття
        для певного тримача
        :param type_holder:
        :return: tuple(angles(), heights(), distance())
        """
        sheet_die = book["Матриця одноручова"]
        rows = sheet_die.max_row
        set_angle = set()
        set_height = set()
        set_distance = set()

        for index in range(2, rows):
            if sheet_die["B" + str(index)].value == type_holder:
                set_angle.add(sheet_die["J" + str(index)].value)
                set_height.add(sheet_die["L" + str(index)].value)
                set_distance.add(sheet_die["K" + str(index)].value)

        if type_holder == "Amada-promecam":
            sheet_die = book["Матриця багаторучова"]
            rows = sheet_die.max_row
            for index in range(2, rows):
                if sheet_die["B" + str(index)].value == type_holder:
                    set_angle.add(sheet_die["J" + str(index)].value)
                    set_angle.add(sheet_die["M" + str(index)].value)
                    set_angle.add(sheet_die["P" + str(index)].value)
                    set_angle.add(sheet_die["S" + str(index)].value)

                    set_height.add(sheet_die["V" + str(index)].value)

                    set_distance.add(sheet_die["K" + str(index)].value)
                    set_distance.add(sheet_die["N" + str(index)].value)
                    set_distance.add(sheet_die["Q" + str(index)].value)
                    set_distance.add(sheet_die["T" + str(index)].value)

        tuple_angle = tuple(sorted(set_angle))
        tuple_height = tuple(sorted(set_height))
        tuple_distance = tuple(sorted(set_distance))
        return (
            tuple_angle,
            tuple_height,
            tuple_distance
        )
