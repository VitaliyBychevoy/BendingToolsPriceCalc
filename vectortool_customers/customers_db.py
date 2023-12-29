from openpyxl import *

#DEFAULT_PATH = "vectortool_customers/customers_vectortool.xlsx"
DEFAULT_PATH = "vectortool_customers/customers.xlsx"
#DEFAULT_PATH = "data/customers_vectortool.xlsx"

company_index: int = 0


def get_short_name_list() -> list:

    short_name_list: list = ["Оберіть компанію"]

    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]

    rows = worksheet.max_row
    short_name_list_item = list()
    for item in range(1, rows + 1):
        short_name_list_item.append(worksheet["A" + str(item)].value)

    wb.close()

    short_name_list.extend(sorted(short_name_list_item))
    return short_name_list

def get_full_name_list() -> list:
    full_name_list: list = []
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]

    rows = worksheet.max_row

    for item in range(1, rows + 1):
        full_name_list.append(worksheet["B" + str(item)].value)

    wb.close()
    return full_name_list


def get_full_name_company(short_name: str) -> str:
    short_name_list: list = get_short_name_list()[1:]
    #full_name_index: int = short_name_list.index(short_name) + 1
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]
    full_name_company = ""
    #full_name_company = worksheet["B" + str(full_name_index)].value
    for item in range(1, worksheet.max_row+1):
        if worksheet["A" + str(item)].value == short_name:
            full_name_company = worksheet["B" + str(item)].value
            break
    wb.close()
    return full_name_company


def up_date_full_name_company(
        short_name: str,
        new_full_name: str
) -> None:
    """
    Функція змінює повне ім'я компанії
    на new_full_name по short_name
    """
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]
    max_row = worksheet.max_row
    for item in range(1, max_row + 1):
        if worksheet["A" + str(item)].value == short_name:
            worksheet["B" + str(item)] = new_full_name
            break
    wb.save(DEFAULT_PATH)
    #wb.save("customers_vectortool.xlsx")
    wb.close()


def up_date_short_name(
        full_name,
        new_short_name
) -> None:
    """
    Функція змінює коротке ім'я компанії по повному імені
    :param full_name: str
    :param new_short_name: str
    :return: None
    """
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]
    max_row = worksheet.max_row
    for item in range(1, max_row + 1):
        if worksheet["B" + str(item)].value == full_name:
            worksheet["A" + str(item)] = new_short_name
            break
    wb.save(DEFAULT_PATH)
    #wb.save("customers_vectortool.xlsx")
    wb.close()

def add_new_company(short_name, full_name) -> None:
    """
    Функція додає нову компаню
    :param short_name: str
    :param full_name: str
    :return:
    """
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]
    max_row = worksheet.max_row
    worksheet["A" + str(max_row + 1)] = short_name
    worksheet["B" + str(max_row + 1)] = full_name
    wb.save(DEFAULT_PATH)
    #wb.save("customers_vectortool.xlsx")
    wb.close()


def delete_customer(short_name: str) -> None:
    """
    Функція видаляє клієнта з бази
    :param short_name: str
    :return:
    """
    wb = load_workbook(DEFAULT_PATH)
    #wb = load_workbook("customers_vectortool.xlsx")
    worksheet = wb[" Companies"]
    max_row = worksheet.max_row
    for item in range(1, max_row + 1):
        if worksheet["A" + str(item)].value == short_name:
            worksheet.delete_rows(item, 1)
            break
    wb.save(DEFAULT_PATH)
    #wb.save("customers_vectortool.xlsx")
    wb.close()

