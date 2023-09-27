from datetime import datetime, date
import openpyxl.styles.numbers

from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import *
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.page import PageMargins
import openpyxl.styles.numbers
from openpyxl.styles import Font, Fill #Стилі для текста
from openpyxl.styles import PatternFill #Cтили для ячеєк

import model
from vectortool_customers.customers_db import *
from model import *

#Шрифти
company_description_font = Font(size=8, italic=True)
company_description_font.name = "Times New Roman"


date_font = Font(size=9,  bold=True)
date_font.name = "Times New Roman"

customer_name_font = Font(size=9,  bold=True)
customer_name_font.name = "Times New Roman"

table_title_font = Font(size=10,  bold=True)
table_title_font.name = "Times New Roman"

table_head_font = Font(size=5,  bold=True)
table_head_font.name = "Times New Roman"

description_font = Font(size=8)
description_font.name = "Arial Narrow"

description_ua_font = Font(size=7)
description_ua_font.name = "Times New Roman"

position_font = Font(size=7, bold=True)
position_font.name = "Times New Roman"

numbers_table_font = Font(size=7)
numbers_table_font.name = "Times New Roman"

total_bill_font = Font(size=8)
total_bill_font.name = "Times New Roman"

discount_font = Font(size=8, bold=True)
discount_font.name = "Times New Roman"

after_table_font = Font(size=7)
after_table_font.name = "Arial Narrow"

italic_bold = Font(size=8, italic=True, bold=True)
italic_bold.name = "Times New Roman"

#Рамка
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

#Центрування
alignment_right_center = Alignment(
        horizontal="right",
        vertical='center',
        wrapText=True
    )

alignment_left_center = Alignment(
        horizontal="left",
        vertical='center',
        wrapText=True
    )

alignment_center_top = Alignment(
        horizontal="center",
        vertical='top',
        wrapText=True
    )

#Кольори
pink_color = PatternFill(
    fill_type='solid',
    start_color='ff1493',
    end_color='ff1493'
)

yellow_color = PatternFill(
    fill_type='solid',
    start_color='ffff00',
    end_color='ffff00'
)
def name_offer(customer_name: str) -> str:
    this_moment_list = str(datetime.now()).split(" ")
    today = this_moment_list[0].split("-")[::-1]
    this_time = this_moment_list[1].split(":")[:2]
    return f"{customer_name}_({this_time[0]} {this_time[1]})_{today[0]}.{today[1]}.{today[2]}.xlsx"


#Встановлюємо ширину та колір стовбчиків
def column_style(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet.column_dimensions['A'].width = 1 * 1.72
    sheet.column_dimensions['B'].width = 1.57 * 1.5
    sheet.column_dimensions['C'].width = 1 * 1.72
    sheet.column_dimensions['D'].width = 24.43
    sheet.column_dimensions['D'].fill = yellow_color
    sheet.column_dimensions['E'].width = 1.14 * 1.72
    sheet.column_dimensions['E'].fill = yellow_color
    sheet.column_dimensions['F'].width = 43.29 * 1.0169
    sheet.column_dimensions['G'].width = 1.43 * 1.5
    sheet.column_dimensions['H'].width = 25.71 * 1.0284
    sheet.column_dimensions['I'].width = 4.43 * 1.72
    sheet.column_dimensions['I'].fill = yellow_color
    sheet.column_dimensions['J'].width = 2.71 * 1.72
    sheet.column_dimensions['J'].fill = yellow_color
    sheet.column_dimensions['K'].width = 3.14 * 1.25
    sheet.column_dimensions['L'].width = 4.71 * 1.72
    sheet.column_dimensions['L'].fill = yellow_color
    sheet.column_dimensions['M'].fill = yellow_color
    sheet.column_dimensions['M'].width = 9 * 1.9
    sheet.column_dimensions['N'].width = 9 * 1.9
    sheet.column_dimensions['N'].fill = yellow_color
    sheet.column_dimensions['O'].width = 9 * 1.08
    sheet.column_dimensions['P'].width = 9 * 1.08
    sheet.column_dimensions['Q'].width = 9 * 1.08
    sheet.column_dimensions['R'].width = 9 * 1.08
    sheet.column_dimensions['S'].width = 9 * 1.2
    sheet.column_dimensions['S'].fill = pink_color
    sheet.column_dimensions['T'].width = 9 * 1.5
    sheet.column_dimensions['T'].fill = pink_color

    sheet.column_dimensions['U'].width = 1 * 1.72


#Оброблюємо строки до таблички
def row_style(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet.row_dimensions[6].height = 15 * 0.3
    sheet.row_dimensions[7].height = 15 * 2.38
    sheet.row_dimensions[8].height = 15 * 4.35
    sheet.row_dimensions[9].height = 15 * 0.24
    sheet.row_dimensions[10].height = 15 * 0.75
    sheet.row_dimensions[11].height = 15 * 0.24
    sheet.row_dimensions[12].height = 15 * 0.75
    sheet.row_dimensions[13].height = 15 * 0.24
    sheet.row_dimensions[14].height = 15 * 0.8
    sheet.row_dimensions[15].height = 15 * 0.45
    sheet.row_dimensions[16].height = 15 * 0.65

#Поєднуємо необхідні комірки до таблиці
def merge_cells_before_table(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet.merge_cells(f'B8:P8')
    sheet.merge_cells(f'O10:R10')
    sheet.merge_cells(f'B12:P12')


def fill_company_info(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet['B8'].font = company_description_font
    sheet['B8'].alignment = Alignment(
        horizontal="center",
        vertical='center',
        wrapText=True
    )
    sheet['B8'].value = \
    "Компанія Tecnostamp (TS) засновано у 1978 году в м." +\
     " П'яченца на півночі Італії і сьогодні є найбільшим " +\
      "світовим виробником листозгинального інструменту. " +\
      "Бренд Tecnostamp " +\
     "представлений більш ніж в 30-ти країнах світу і займає " +\
      "провідні позиції на ринках Німеччини, Росії, Італії та США. " +\
     "Підбір інструменту за кресленнями та технічні консультації. " +\
     "Проектування спеціального інструменту під нестандартні " +\
     "задачі.  Обираючи бренд Tecnostamp, Ви отримуєте те " +\
     "надійний інструмент для гнуття преміум класу з великим " +\
     "терміном служби і гарантією бездоганної якості. Інструмент " +\
     "для гнуття для наступних верстатів:AMADA; TRUMPF; MVD; " +\
     "INAN;PRIMA POWER; FINN POWER; LVD;Bystronic; " +\
     "Safan; Salvagnini;EHT; Boschert; Darley;Gasparini; HACO; " +\
     "Farina; Schiavi; Adira; Guifil; Jordi; Ursviken; Hammerle; " +\
     "Dener; Durma; Ermaksan; Baykal."


def fill_today_before_table(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    #Загальний формат комірки змінюємо на формат дату
    sheet['F10'].number_format = \
        openpyxl.styles.numbers.BUILTIN_FORMATS[14]
    sheet['F10'] = "=TODAY()"
    sheet['F10'].font = date_font
    sheet['F10'].alignment = Alignment(
        horizontal="left",
        vertical='top',
        wrapText=True
    )


#Заповнюємо назву компанії
def fill_customer_name(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        customer_name: str
) -> None:
    sheet["O10"].font = customer_name_font
    sheet['O10'].alignment = Alignment(
        horizontal="right",
        vertical='top',
        wrapText=True
    )
    sheet["O10"].value = get_full_name_company(customer_name)

#Заповнюємо назву таблиці
def fill_title_table(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet['B12'] = \
    "Техніко-комерційна пропозиція на постачання інструменту " +\
    "TECNOSTAMP S.R.L для листозгинального пресу з ЧПК"
    sheet['B12'] .font =  table_title_font


    sheet['B12'].alignment = Alignment(
    horizontal="center",
    vertical='center',
    wrapText=True
    )

#Заповнюємо назву таблиці
def fill_table_head(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet['B14'].value = "№"
    sheet['B14'].font = table_head_font
    sheet['B14'].border = thin_border
    sheet['B14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['C14'].border = thin_border

    sheet['D14'].value = "Description"
    sheet['D14'].font = table_head_font
    sheet['D14'].border = thin_border
    sheet['D14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['D14'].fill = yellow_color

    sheet['E14'].border = thin_border
    sheet['E14'].fill = yellow_color

    sheet['F14'].value = "Опис"
    sheet['F14'].font = table_head_font
    sheet['F14'].border = thin_border
    sheet['F14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['G14'].value = "/"
    sheet['G14'].font = table_head_font
    sheet['G14'].border = thin_border
    sheet['G14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['H14'].value = "Розмір, мм"
    sheet['H14'].font = table_head_font
    sheet['H14'].border = thin_border
    sheet['H14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['I14'].border = thin_border
    sheet['I14'].fill = yellow_color

    sheet['J14'].value = "Вага"
    sheet['J14'].font = table_head_font
    sheet['J14'].border = thin_border
    sheet['J14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['J14'].fill =yellow_color

    sheet['K14'].value = "Кіл-ть"
    sheet['K14'].font = table_head_font
    sheet['K14'].border = thin_border
    sheet['K14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['L14'].value = "ЗАКУПКА"
    sheet['L14'].font = table_head_font
    sheet['L14'].border = thin_border
    sheet['L14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['L14'].fill = yellow_color

    sheet['M14'].value = "Вартість позиції"
    sheet['M14'].font = table_head_font
    sheet['M14'].border = thin_border
    sheet['M14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['M14'].fill = yellow_color

    sheet['N14'].value = "Відсоток від вартості позиції"
    sheet['N14'].font = table_head_font
    sheet['N14'].border = thin_border
    sheet['N14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['N14'].fill = yellow_color

    sheet['O14'].value = "Ціна од. EURO"
    sheet['O14'].font = table_head_font
    sheet['O14'].border = thin_border
    sheet['O14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['P14'].value = "Ціна разом EURO"
    sheet['P14'].font = table_head_font
    sheet['P14'].border = thin_border
    sheet['P14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['Q14'].value = "Ціна од. ГРН"
    sheet['Q14'].font = table_head_font
    sheet['Q14'].border = thin_border
    sheet['Q14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['R14'].value = "Ціна разом ГРН"
    sheet['R14'].font = table_head_font
    sheet['R14'].border = thin_border
    sheet['R14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['S14'].value = "1C за одиницю UAH"
    sheet['S14'].font = table_head_font
    sheet['S14'].border = thin_border
    sheet['S14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['S14'].fill = pink_color

    sheet['T14'].value = "1C разом UAH"
    sheet['T14'].font = table_head_font
    sheet['T14'].border = thin_border
    sheet['T14'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['T14'].fill = yellow_color

def fill_number_string(
        sheet: openpyxl.worksheet.worksheet.Worksheet
) -> None:
    sheet['B15'].value = 1
    sheet['B15'].font = table_head_font
    sheet['B15'].border = thin_border
    sheet['B15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['C15'].border = thin_border
    sheet['C15'].value = 2
    sheet['C15'].font = table_head_font
    sheet['C15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['D15'].font = table_head_font
    sheet['D15'].border = thin_border
    sheet['D15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )
    sheet['D15'].fill = yellow_color

    sheet['E15'].border = thin_border
    sheet['E15'].fill = yellow_color

    sheet['F15'].value = 3
    sheet['F15'].font = table_head_font
    sheet['F15'].border = thin_border
    sheet['F15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['G15'].font = table_head_font
    sheet['G15'].border = thin_border
    sheet['G15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['H15'].value = 4
    sheet['H15'].font = table_head_font
    sheet['H15'].border = thin_border
    sheet['H15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['I15'].border = thin_border
    sheet['I15'].fill = yellow_color


    sheet['J15'].font = table_head_font
    sheet['J15'].border = thin_border
    sheet['J15'].fill = yellow_color

    sheet['K15'].value = 5
    sheet['K15'].font = table_head_font
    sheet['K15'].border = thin_border
    sheet['K15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['L15'].font = table_head_font
    sheet['L15'].border = thin_border
    sheet['L15'].fill = yellow_color

    sheet['M15'].font = table_head_font
    sheet['M15'].border = thin_border
    sheet['M15'].fill = yellow_color

    sheet['N15'].font = table_head_font
    sheet['N15'].border = thin_border
    sheet['N15'].fill = yellow_color

    sheet['O15'].value = 6
    sheet['O15'].font = table_head_font
    sheet['O15'].border = thin_border
    sheet['O15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['P15'].value = 7
    sheet['P15'].font = table_head_font
    sheet['P15'].border = thin_border
    sheet['P15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['Q15'].value = 8
    sheet['Q15'].font = table_head_font
    sheet['Q15'].border = thin_border
    sheet['Q15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['R15'].value = 9
    sheet['R15'].font = table_head_font
    sheet['R15'].border = thin_border
    sheet['R15'].alignment = Alignment(
        horizontal="center",
        vertical='center'
    )

    sheet['S15'].border = thin_border
    sheet['S15'].fill = pink_color

    sheet['T15'].border = thin_border
    sheet['T15'].fill = yellow_color

def empty_string(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        number_string: int
) -> None:

    sheet[f'B{str(number_string)}'].border = thin_border

    sheet[f'C{str(number_string)}'].border = thin_border

    sheet[f'D{str(number_string)}'].border = thin_border
    sheet[f'D{str(number_string)}'].fill = PatternFill(
        fill_type='solid',
        start_color='ffff00',
        end_color='ffff00'
    )

    sheet[f'E{str(number_string)}'].border = thin_border
    sheet[f'E{str(number_string)}'].fill = yellow_color

    sheet[f'F{str(number_string)}'].border = thin_border

    sheet[f'G{str(number_string)}'].border = thin_border

    sheet[f'H{str(number_string)}'].border = thin_border



    sheet[f'I{str(number_string)}'].border = thin_border
    sheet[f'I{str(number_string)}'].fill = yellow_color

    sheet[f'J{str(number_string)}'].border = thin_border
    sheet[f'J{str(number_string)}'].fill = yellow_color

    sheet[f'K{str(number_string)}'].border = thin_border

    sheet[f'L{str(number_string)}'].border = thin_border
    sheet[f'L{str(number_string)}'].fill = yellow_color

    sheet[f'M{str(number_string)}'].border = thin_border
    sheet[f'M{str(number_string)}'].fill = yellow_color

    sheet[f'N{str(number_string)}'].border = thin_border
    sheet[f'N{str(number_string)}'].fill = yellow_color

    sheet[f'O{str(number_string)}'].border = thin_border

    sheet[f'P{str(number_string)}'].border = thin_border

    sheet[f'Q{str(number_string)}'].border = thin_border

    sheet[f'R{str(number_string)}'].border = thin_border

    sheet[f'S{str(number_string)}'].border = thin_border
    sheet[f'S{str(number_string)}'].fill = pink_color

    sheet[f'T{str(number_string)}'].border = thin_border
    sheet[f'T{str(number_string)}'].fill = yellow_color
def items_in_row(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        invoice: Invoice,
        current_row: int
) -> int:

    for index in range(len(invoice.get_list_item())):

        sheet.row_dimensions[current_row].height = 150

        sheet[f"B{str(current_row)}"].value = index + 1
        sheet[f"B{str(current_row)}"].border = thin_border
        sheet[f"B{str(current_row)}"].alignment = \
            Alignment(horizontal="center", vertical='center')
        sheet[f"B{str(current_row)}"].font = position_font

        sheet[f'C{str(current_row)}'].border = thin_border

        sheet[f"D{str(current_row)}"].border = thin_border
        sheet[f"D{str(current_row)}"].value = \
            invoice.get_list_item()[index].get_en_name_item()
        sheet[f"D{str(current_row)}"].font = description_font
        sheet[f"D{str(current_row)}"].alignment = Alignment(
            horizontal="left",
            vertical='center',
            wrapText=True
        )
        sheet[f'D{str(current_row)}'].fill = yellow_color

        sheet[f"E{str(current_row)}"].border = thin_border
        sheet[f'E{str(current_row)}'].fill = yellow_color

        sheet[f"F{str(current_row)}"].border = thin_border
        sheet[f"F{str(current_row)}"].value = \
            invoice.get_list_item()[index].get_ua_name_item()
        sheet[f"F{str(current_row)}"].font = description_ua_font
        sheet[f"F{str(current_row)}"].alignment = Alignment(
            horizontal="left",
            vertical='center',
            wrapText=True
        )

        sheet[f"G{str(current_row)}"].border = thin_border

        sheet[f"H{str(current_row)}"].border = thin_border
        img = openpyxl.drawing.image.Image(f"data/{invoice.get_list_item()[index].get_image_path()}")
        img.height = 160
        img.width = 140
        img.anchor = f"H{str(current_row)}"
        sheet.add_image(img)

        sheet[f"K{str(current_row)}"].border = thin_border
        sheet[f"K{str(current_row)}"].value = invoice.get_list_item()[index].get_amount_item()
        sheet[f"K{str(current_row)}"].font = numbers_table_font
        sheet[f"K{str(current_row)}"].alignment = alignment_right_center

        sheet[f"J{str(current_row)}"].border = thin_border
        sheet[f"J{str(current_row)}"].font = numbers_table_font
        sheet[f"J{str(current_row)}"].alignment = alignment_right_center
        sheet[f'J{str(current_row)}'].fill = yellow_color
        sheet[f"J{str(current_row)}"].value = \
            invoice.get_list_item()[index].get_weight_item()

        sheet[f"I{str(current_row)}"].border = thin_border
        sheet[f"I{str(current_row)}"].font = numbers_table_font
        sheet[f"I{str(current_row)}"].alignment = alignment_right_center
        sheet[f'I{str(current_row)}'].fill = yellow_color
        sheet[f"I{str(current_row)}"].value = \
            f"=J{str(current_row)}*K{str(current_row)}"

        sheet[f"L{str(current_row)}"].border = thin_border
        sheet[f"L{str(current_row)}"].font = numbers_table_font
        sheet[f"L{str(current_row)}"].value = \
            f"={invoice.get_list_item()[index].get_price_item()}*((100-{float(invoice.get_provider_discount())})/100)"
        sheet[f'L{str(current_row)}'].fill = yellow_color
        sheet[f"L{str(current_row)}"].alignment = alignment_right_center

        sheet[f"M{str(current_row)}"].border = thin_border
        sheet[f"M{str(current_row)}"].font = numbers_table_font
        sheet[f"M{str(current_row)}"].value = f"=L{str(current_row)}*K{str(current_row)}"
        sheet[f'M{str(current_row)}'].fill = yellow_color
        sheet[f"M{str(current_row)}"].alignment = alignment_right_center
        sheet[f"N{str(current_row)}"].border = thin_border
        sheet[f"N{str(current_row)}"].font = numbers_table_font
        sheet[f"N{str(current_row)}"].value = f"=((L{str(current_row)}*100)/{invoice.get_sum_item_price()})/100"
        sheet[f"N{str(current_row)}"].number_format = '#,##0.00'
        sheet[f'N{str(current_row)}'].fill = yellow_color
        sheet[f"N{str(current_row)}"].alignment = alignment_right_center

        sheet[f"O{str(current_row)}"].value = f"=Q{str(current_row)}/{str(invoice.get_rate())}"
        sheet[f"O{str(current_row)}"].number_format = '#,##0.00'
        sheet[f"O{str(current_row)}"].border = thin_border
        sheet[f"O{str(current_row)}"].font = numbers_table_font
        sheet[f"O{str(current_row)}"].alignment = Alignment(
            horizontal="center",
            vertical='center'
        )

        sheet[f"P{str(current_row)}"].value = f"=O{str(current_row)}*K{str(current_row)}"
        sheet[f"P{str(current_row)}"].number_format = '#,##0.00'
        sheet[f"P{str(current_row)}"].border = thin_border
        sheet[f"P{str(current_row)}"].font = numbers_table_font
        sheet[f"P{str(current_row)}"].alignment = alignment_right_center

        sheet[f"R{str(current_row)}"].value = f"={invoice.get_total_price_ua()}*K{str(current_row)}*N{str(current_row)}"
        sheet[f"R{str(current_row)}"].number_format = '#,##0.00'
        sheet[f"R{str(current_row)}"].border = thin_border
        sheet[f"R{str(current_row)}"].font = numbers_table_font
        sheet[f"R{str(current_row)}"].alignment = alignment_right_center

        sheet[f"Q{str(current_row)}"].value = f"=R{str(current_row)}/K{str(current_row)}"
        sheet[f"Q{str(current_row)}"].number_format = '#,##0.00'
        sheet[f"Q{str(current_row)}"].border = thin_border
        sheet[f"Q{str(current_row)}"].font = numbers_table_font
        sheet[f"Q{str(current_row)}"].alignment = alignment_right_center

        current_row += 1

    print("Current row: ", current_row, ".")
    return current_row

def write_row(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        item: Item,
        number_string: int,
        index: int,
        provider_discount: float,
        rate: str
) -> None:
    sheet[f"B{str(number_string)}"].value = index + 1
    sheet[f"B{str(number_string)}"].border = thin_border
    sheet[f"B{str(number_string)}"].alignment = \
        Alignment(horizontal="center", vertical='center')
    sheet[f"B{str(number_string)}"].font = position_font

    sheet[f"C{str(number_string)}"].border = thin_border

    sheet[f"D{str(number_string)}"].border = thin_border
    sheet[f"D{str(number_string)}"].value = \
        item.get_en_name_item()
    sheet[f"D{str(number_string)}"].font = description_font
    sheet[f"D{str(number_string)}"].alignment = Alignment(
        horizontal="left",
        vertical='center',
        wrapText=True
    )
    sheet[f'D{str(number_string)}'].fill = yellow_color

    sheet[f"E{str(number_string)}"].border = thin_border
    sheet[f'E{str(number_string)}'].fill = yellow_color

    sheet[f"F{str(number_string)}"].border = thin_border
    sheet[f"F{str(number_string)}"].value = \
        item.get_ua_name_item()
    sheet[f"F{str(number_string)}"].font = description_ua_font
    sheet[f"F{str(number_string)}"].alignment = Alignment(
        horizontal="left",
        vertical='center',
        wrapText=True
    )

    sheet[f"G{str(number_string)}"].border = thin_border

    sheet[f"H{str(number_string)}"].border = thin_border
    img = openpyxl.drawing.image.Image(f"data/{item.get_image_path()}")
    img.height = 160
    img.width = 140
    img.anchor = f"H{str(number_string)}"
    sheet.add_image(img)

    sheet[f"K{str(number_string)}"].border = thin_border
    sheet[f"K{str(number_string)}"].value = item.get_amount_item()
    sheet[f"K{str(number_string)}"].font = numbers_table_font
    sheet[f"K{str(number_string)}"].alignment = alignment_right_center

    sheet[f"J{str(number_string)}"].border = thin_border
    sheet[f"J{str(number_string)}"].font = numbers_table_font
    sheet[f"J{str(number_string)}"].alignment = alignment_right_center
    sheet[f'J{str(number_string)}'].fill = yellow_color
    sheet[f"J{str(number_string)}"].value = \
        item.get_weight_item()

    sheet[f"I{str(number_string)}"].border = thin_border
    sheet[f"I{str(number_string)}"].font = numbers_table_font
    sheet[f"I{str(number_string)}"].alignment = alignment_right_center
    sheet[f'I{str(number_string)}'].fill = yellow_color

    sheet[f"I{str(number_string)}"].value = \
        f"=J{str(number_string)}*K{str(number_string)}"

    sheet[f"L{str(number_string)}"].border = thin_border
    sheet[f"L{str(number_string)}"].font = numbers_table_font
    sheet[f"L{str(number_string)}"].value = \
        f"={item.get_price_item()}*((100-{provider_discount})/100)"
    sheet[f'L{str(number_string)}'].fill = yellow_color

    sheet[f"L{str(number_string)}"].alignment = alignment_right_center

    sheet[f"O{str(number_string)}"].border = thin_border
    sheet[f"O{str(number_string)}"].value = 0
    sheet[f"O{str(number_string)}"].font = numbers_table_font
    sheet[f"O{str(number_string)}"].alignment = alignment_right_center

    sheet[f"M{str(number_string)}"].border = thin_border
    sheet[f"M{str(number_string)}"].value = \
        f"=O{str(number_string)}/{float(str(rate).replace(',','.'))}"
    sheet[f"M{str(number_string)}"].font = numbers_table_font
    sheet[f"M{str(number_string)}"].alignment = alignment_right_center

    sheet[f"M{str(number_string)}"].number_format = '#,##0.00'

    sheet[f"N{str(number_string)}"].border = thin_border
    sheet[f"N{str(number_string)}"].value = \
        f"= M{str(number_string)} * K{str(number_string)}"
    sheet[f"N{str(number_string)}"].font = numbers_table_font
    sheet[f"N{str(number_string)}"].alignment = alignment_right_center
    sheet[f"N{str(number_string)}"].number_format = '#,##0.00'

    sheet[f"P{str(number_string)}"].border = thin_border
    sheet[f"P{str(number_string)}"].value = \
        f"= O{str(number_string)} * K{str(number_string)}"
    sheet[f"P{str(number_string)}"].font = numbers_table_font
    sheet[f"P{str(number_string)}"].alignment = alignment_right_center


def total_weight(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int) -> None:
        sheet[f"I{str(current_row)}"].value = f"=SUM(I17:I{str(current_row-1)})"
        sheet[f"I{str(current_row)}"].alignment = alignment_right_center


def fill_last_row_table(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        amount_items: int,
        current_row: int
) -> None:
    empty_string(
        sheet,
        current_row
    )
    if amount_items < 2:
        sheet[f"I{str(current_row)}"] = \
            f"=SUM(I17:I17)"
    else:
        sheet[f"I{str(current_row)}"] = \
            total_weight(current_row - 1)
        sheet[f"I{str(current_row)}"].alignment = alignment_right_center
        sheet[f"I{str(current_row)}"].font = numbers_table_font
        sheet[f"I{str(current_row)}"].fill =yellow_color

def fill_total_bill(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f"F{str(current_row)}"].font = total_bill_font
    sheet[f"F{str(current_row)}"].value = "Разом"
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=SUM(P17:P{str(current_row-2)})"
    sheet[f"P{str(current_row)}"].font = numbers_table_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"P{str(current_row)}"].alignment = alignment_right_center

    sheet[f"Q{str(current_row)}"].font = numbers_table_font
    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=SUM(R17:R{str(current_row-2)})"
    sheet[f"R{str(current_row)}"].font = numbers_table_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"R{str(current_row)}"].alignment = alignment_right_center

    sheet[f"T{str(current_row)}"] = \
        f"=SUM(T17:T{str(current_row-2)})"
    sheet[f"T{str(current_row)}"].font = numbers_table_font
    sheet[f"T{str(current_row)}"].border = thin_border
    sheet[f"T{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"T{str(current_row)}"].alignment = alignment_right_center
    sheet[f"T{str(current_row)}"].fill = pink_color

def get_price_item_for_customer(invoce: Invoice) -> list:
    #Загальна ціна покупки (ціна * кількість)

    total_price: float = sum([
        invoce.get_list_item()[index].get_amount_item() *
        ((100-float(invoce.get_provider_discount()))/100) *
        invoce.get_list_item()[index].get_price_item()
        for index in range(len(invoce.get_list_item()))
    ])

    print("Total price: ", total_price, " ", type(total_price))
    print("get price: ",
          invoce.get_packing_price(), " ", type(invoce.get_packing_price()))

    total_price = total_price + float(invoce.get_packing_price().replace(",","."))

    print("Total price (+packing): ", total_price, " ", type(total_price))

    total_price = round(total_price * model.BANK_TAX, 2)

    print("Total price(+BANK_TAX): ", total_price, " ", type(total_price))

    total_price = round((total_price /
                   ((100 - float(invoce.get_commission_percentage().replace(",","."))) / 100)), 2)
    print("Total price(+Commission): ", total_price, " ", type(total_price))

    total_price = round(total_price * float(invoce.get_rate().replace(",", ".")), 2)
    print("Total price(UA): ", total_price, " ", type(total_price), " грн.")

    prices_item = [item.get_price_item() for item in invoce.get_list_item()]
    print(prices_item)
    sum_prices_item = sum(prices_item)
    price_item_ua =[]
    for index in range(len(invoce.get_list_item())):
        price_item_ua.append(round(
            ((100 * invoce.get_list_item()[index].get_price_item() / sum_prices_item) * total_price / 100),  2))

    print("price_item_ua ", price_item_ua)
    print("sum(price_item_ua: )", sum(price_item_ua))
    return price_item_ua

def tax_row_total(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int
) -> None:

    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f"F{str(current_row)}"].font = total_bill_font
    sheet[f"F{str(current_row)}"].value = "ПДВ"
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-1)}*0.2"
    sheet[f"P{str(current_row)}"].font = numbers_table_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"P{str(current_row)}"].alignment =  alignment_right_center

    sheet[f"Q{str(current_row)}"].font = numbers_table_font
    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=R{str(current_row-1)}*0.2"
    sheet[f"R{str(current_row)}"].font = numbers_table_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"R{str(current_row)}"].alignment =  alignment_right_center

    sheet[f"T{str(current_row)}"] = f"=T{str(current_row-1)}*0.2"
    sheet[f"T{str(current_row)}"].border = thin_border
    sheet[f"T{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"T{str(current_row)}"].fill = pink_color
    sheet[f"T{str(current_row)}"].alignment = alignment_right_center
    sheet[f"T{str(current_row)}"].font = numbers_table_font

def set_price(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        list_price: list
) -> None:
    start_row: int = 17
    for index in range(len(list_price)):
        sheet[f"O{start_row}"].value = list_price[index]
        start_row += 1


def fill_total_tax(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:M{str(current_row)}')
    sheet[f"F{str(current_row)}"].font = total_bill_font
    sheet[f"F{str(current_row)}"].value = \
        "Загальна вартість з ПДВ"
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border

    sheet[f"N{str(current_row)}"] = \
        f"=N{str(current_row-1)}+N{str(current_row-2)}"
    sheet[f"N{str(current_row)}"].font = numbers_table_font
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].number_format = '#,##0.00'


    sheet[f"O{str(current_row)}"].font = numbers_table_font
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-1)}+P{str(current_row-2)}"
    sheet[f"P{str(current_row)}"].font = numbers_table_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'


def fill_discount_customer_value(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
        customer_name: str,
        customer_discount: str
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f'F{str(current_row)}'] =\
         f"Знижка для компанії" + \
         f" {get_full_name_company(customer_name)}" + \
         f"  {customer_discount}%"
    sheet[f'F{str(current_row)}'].font = discount_font
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-1)}*({float(customer_discount)}/100)"
    sheet[f"P{str(current_row)}"].font = discount_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'

    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=R{str(current_row-1)}*({float(customer_discount)}/100)"
    sheet[f"R{str(current_row)}"].font = discount_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'


def fill_total_tax_discount(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f'F{str(current_row)}'] =\
         f"Вартість з урахуванням знижки"

    sheet[f'F{str(current_row)}'].font = discount_font
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-2)}-P{str(current_row-1)}"
    sheet[f"P{str(current_row)}"].font = discount_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"P{str(current_row)}"].font = discount_font

    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=R{str(current_row-2)}-R{str(current_row-1)}"
    sheet[f"R{str(current_row)}"].font = discount_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"R{str(current_row)}"].font = discount_font


def fill_delivery_value(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
        invoice: Invoice
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f'F{str(current_row)}'] =\
         f"Вартість доставки до складу у місті Київ"

    sheet[f'F{str(current_row)}'].font = discount_font
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"={invoice.get_total_delivery_price_ua()}"
    sheet[f"R{str(current_row)}"].font = discount_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'

    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=R{str(current_row)}/{invoice.get_rate()}"
    sheet[f"P{str(current_row)}"].font = discount_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'

def total_bill_with_tax(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f"F{str(current_row)}"].font = total_bill_font
    sheet[f"F{str(current_row)}"].value = "Вартість разом з ПДВ"
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-1)}+P{str(current_row-2)}"
    sheet[f"P{str(current_row)}"].font = numbers_table_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"P{str(current_row)}"].alignment = alignment_right_center

    sheet[f"Q{str(current_row)}"].font = numbers_table_font
    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=R{str(current_row-1)}+R{str(current_row-2)}"
    sheet[f"R{str(current_row)}"].font = numbers_table_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"R{str(current_row)}"].alignment = alignment_right_center

    sheet[f"T{str(current_row)}"] = \
        f"=T{str(current_row-1)}+T{str(current_row-2)}"
    sheet[f"T{str(current_row)}"].font = numbers_table_font
    sheet[f"T{str(current_row)}"].border = thin_border
    sheet[f"T{str(current_row)}"].number_format = '#,##0.00'
    sheet[f"T{str(current_row)}"].fill = pink_color
    sheet[f"T{str(current_row)}"].alignment = alignment_right_center
    sheet[f"T{str(current_row)}"].font = numbers_table_font

def fill_total_price(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int
) -> None:
    sheet.merge_cells(f'F{str(current_row)}:O{str(current_row)}')
    sheet[f"F{str(current_row)}"].font = discount_font
    sheet[f"F{str(current_row)}"].value = "Загальна вартість"
    sheet[f"F{str(current_row)}"].border = thin_border
    sheet[f"G{str(current_row)}"].border = thin_border
    sheet[f"H{str(current_row)}"].border = thin_border
    sheet[f"I{str(current_row)}"].border = thin_border
    sheet[f"J{str(current_row)}"].border = thin_border
    sheet[f"K{str(current_row)}"].border = thin_border
    sheet[f"L{str(current_row)}"].border = thin_border
    sheet[f"M{str(current_row)}"].border = thin_border
    sheet[f"N{str(current_row)}"].border = thin_border
    sheet[f"O{str(current_row)}"].border = thin_border

    sheet[f"P{str(current_row)}"] = \
        f"=P{str(current_row-1)}+P{str(current_row-2)}"
    sheet[f"P{str(current_row)}"].font = discount_font
    sheet[f"P{str(current_row)}"].border = thin_border
    sheet[f"P{str(current_row)}"].number_format = '#,##0.00'

    sheet[f"Q{str(current_row)}"].font = numbers_table_font
    sheet[f"Q{str(current_row)}"].border = thin_border

    sheet[f"R{str(current_row)}"] = \
        f"=R{str(current_row-1)}+R{str(current_row-2)}"
    sheet[f"R{str(current_row)}"].font = discount_font
    sheet[f"R{str(current_row)}"].border = thin_border
    sheet[f"R{str(current_row)}"].number_format = '#,##0.00'


def fill_1C_all(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        invoice: Invoice,
        current_row: int
) -> None:
    row = 17
    for index in range(len(invoice.get_list_item())):
        sheet[f"T{str(row)}"] = f"=R{str(current_row)}*(5/6)*N{str(row)}*K{str(row)}"
        sheet[f"T{str(row)}"].border = thin_border
        sheet[f"T{str(row)}"].number_format = '#,##0.00'
        sheet[f"T{str(row)}"].fill = yellow_color
        sheet[f"T{str(row)}"].alignment = alignment_right_center
        sheet[f"T{str(row)}"].font = numbers_table_font


        sheet[f"S{str(row)}"] = f"=T{str(row)}/K{str(row)}"
        sheet[f"S{str(row)}"].border = thin_border
        sheet[f"S{str(row)}"].number_format = '#,##0.00'
        sheet[f"S{str(row)}"].fill = pink_color
        sheet[f"S{str(row)}"].alignment = alignment_right_center
        sheet[f"S{str(row)}"].font = numbers_table_font

        row += 1


def empty_columns(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
        invoice: Invoice
) -> None:
    if invoice.get_customer_discount() == "0":
        sheet.merge_cells(f'B{str(current_row-4)}:B{str(current_row)}')
        sheet.merge_cells(f'C{str(current_row-4)}:C{str(current_row)}')
        for i in range(5, -1, -1):
            sheet[f"B{str(current_row-i)}"].border = thin_border
        for i in range(5, -1, -1):
            sheet[f"C{str(current_row-i)}"].border = thin_border
    else:
        sheet.merge_cells(f'B{str(current_row-6)}:B{str(current_row)}')
        sheet.merge_cells(f'C{str(current_row-6)}:C{str(current_row)}')
        for i in range(7, -1, -1):
            sheet[f"B{str(current_row-i)}"].border = thin_border
        for i in range(7, -1, -1):
            sheet[f"C{str(current_row-i)}"].border = thin_border


def after_table(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        current_row: int,
        invoice: Invoice
) -> None:
    sheet.row_dimensions[current_row].height = 13 * 0.76
    current_row += 1

    sheet.row_dimensions[current_row].height = 13 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value  = \
        "1. Умови оплати згідно з договором."
    sheet[f'B{str(current_row)}'].font = after_table_font
    sheet[f'B{str(current_row)}'].alignment = alignment_left_center
    current_row += 1

    sheet.row_dimensions[current_row].height = 13 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value  = \
        "2. Термін доставки"
    sheet[f'B{str(current_row)}'].font = after_table_font
    sheet[f'B{str(current_row)}'].alignment = alignment_left_center
    current_row += 1

    sheet.row_dimensions[current_row].height = 13 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value  = \
        ("3. Відвантаження зі складу в м. Київ відбувається "
         "після отримання повної суми оплати, протягом доби, "
         "якщо інші умови не визначено договором.")
    sheet[f'B{str(current_row)}'].font = after_table_font
    sheet[f'B{str(current_row)}'].alignment = alignment_left_center
    current_row += 1

    sheet.row_dimensions[current_row].height = 29 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value  = \
        ('4. Доставка відбувається по всій території Україні '
         'логістичною компанією "Нова пошта" за тарифами '
         'перевізника. \n*Самовивіз зі складу в м. Київ вул. Польова 24.')
    sheet[f'B{str(current_row)}'].font = after_table_font
    sheet[f'B{str(current_row)}'].alignment = alignment_left_center
    current_row += 1

    sheet.row_dimensions[current_row].height = 30 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value  = \
        ('5. Термін дії техніко-комерційної пропозиції 3'
         ' (три) календарних дні.\n*Вартість інструменту може '
         'бути змінено відповідно до змін курсу валют на'
         ' Міжбанку України.')
    sheet[f'B{str(current_row)}'].font = after_table_font
    sheet[f'B{str(current_row)}'].alignment = alignment_left_center
    current_row += 1

    sheet.row_dimensions[current_row].height = 12 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 13 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:F{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = "З повагою,"
    current_row += 1
    sheet.row_dimensions[current_row].height = 14 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:F{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = "Бичевий Віталій"
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:F{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = \
        'Інженер-технолог \nТОВ "ВЕКТОРТУЛ"'
    current_row += 1
    sheet.row_dimensions[current_row].height = 6 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 7 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 19 * 0.76
    current_row += 1

    sheet.row_dimensions[current_row].height = 17 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = \
        (f'Ми високо цінуємо  спільну роботу з '
         f'компанією {invoice.get_customer_name()},'
         f'прагнемо до задоволення ваших виробничих потреб.')
    sheet[f'B{str(current_row)}'].font = description_ua_font
    sheet[f'B{str(current_row)}'].alignment = alignment_center_top
    current_row += 1

    sheet.row_dimensions[current_row].height = 17 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = \
        ("Висловлюю надію на продовження "
         "успішної співпраці на благо наших спільних "
         "інтересів, а також на подальше збільшення "
         "досягнутих показників спільної роботи.")
    sheet[f'B{str(current_row)}'].font = description_ua_font
    sheet[f'B{str(current_row)}'].alignment = alignment_center_top
    current_row += 1

    sheet.row_dimensions[current_row].height = 4 * 0.76
    current_row += 1

    sheet.row_dimensions[current_row].height = 17 * 0.76
    sheet.merge_cells(f'B{str(current_row)}:P{str(current_row)}')
    sheet[f'B{str(current_row)}'].value = \
        "Бажаю  Вам  і компанії успіху і процвітання!"
    sheet[f'B{str(current_row)}'].alignment = alignment_center_top
    sheet[f'B{str(current_row)}'].font = italic_bold
    current_row += 1

    sheet.row_dimensions[current_row].height = 2 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 21 * 0.76
    current_row += 1
    sheet.row_dimensions[current_row].height = 24 * 0.76
    current_row += 1


