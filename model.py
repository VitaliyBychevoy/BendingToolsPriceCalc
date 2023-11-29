import openpyxl
from openpyxl import *
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl.styles.numbers
from openpyxl.styles import Font, Fill #Стилі для текста
from openpyxl.styles import PatternFill #Cтили для ячеєк
from openpyxl.styles import colors #Kольори для текста и ячеєк
from vectortool_customers.customers_db import *
DB_FOLDER = "data/DB_bending.xlsx"


#Bank tax
BANK_TAX: float = 1.002

#Transfer_currency
TRANSFER_CURRENCY: int = 1100

#брокер
BROKER: int = 2500

class Item:

    def __init__(
            self,
            db_path: str = None,
            type_holder: str = None,
            type_item: str = None,
            code_item: str = None,
            length_item: str = None,
            length_item_mm: str = None,
            weight: float = 0.0,
            price_item: float = 0.0,
            price_for_customer_ua: float = 0.0,
            discount_item: float = 0.0,
            cross_section_param: dict = None,
            image_path: str = None,
            en_name_item: str = None,
            ua_name_item: str = None,
            amount_item: int = 0,

    ) -> None:
        self.db = openpyxl.load_workbook(DB_FOLDER)
        self.db_path: str = db_path
        self.type_holder: str = type_holder
        self.type_item: str = type_item
        self.code_item: str = code_item
        self.length_item: str = length_item
        self.length_item_mm: str = length_item_mm
        self.weight: float = weight
        self.price_item: float = price_item
        self.price_for_customer_ua = price_for_customer_ua
        self.discount_item: float = discount_item
        self.cross_section_param: dict = cross_section_param
        self.image_path: str = image_path
        self.en_name_item: str = en_name_item
        self.ua_name_item: str = ua_name_item
        self.amount_item: int = amount_item

    def set_db_path(self, new_path: str) -> None:
        self.db_path = new_path

    def get_db_path(self) -> str:
        return self.db_path

    def set_type_holder(self, new_type_holder: str) -> None:
        self.type_holder = new_type_holder

    def get_type_holder(self) -> str:
        return self.type_holder

    def set_type_item(self, new_type_item: str) -> None:
        self.type_item = new_type_item

    def get_type_item(self) -> str:
        return self.type_item

    def set_code_item(self, new_code_item: str) -> None:
        self.code_item = new_code_item

    def get_code_item(self) -> str:
        return self.code_item

    def set_length_item(self, new_length_item: str) -> None:
        self.length_item = new_length_item

    def get_length_item(self) -> str:
        return self.length_item

    def set_length_item_mm(self, new_length_item_mm: str) -> None:
        self.length_item_mm = new_length_item_mm

    def get_length_item_mm(self) -> str:
        return self.length_item_mm

    def set_weight_item(self, new_weight: float) -> None:
        self.weight = new_weight

    def get_weight_item(self) -> float:
        return self.weight

    def set_price_item(self, new_price_item: float) -> None:
        self.price_item = new_price_item

    def get_price_item(self) -> float:
        return self.price_item

    def set_price_for_customer_ua(self,  price: float) -> None:
        self.price_for_customer_ua = price

    def price_for_customer_ua(self) -> float:
        return self.price_for_customer_ua

    def set_discount_item(self, percent: float) -> None:
        self.discount_item = percent

    def get_discount_item(self) -> float:
        return self.discount_item

    def set_cross_section_param(self, new_dict: dict) -> None:
        self.cross_section_param = new_dict

    def get_cross_section_param(self) -> dict:
        return self.cross_section_param

    def set_image_path(self, new_img_path: str) -> None:
        self.image_path = new_img_path

    def get_image_path(self) -> str:
        return self.image_path

    def set_en_name_item(self, new_en_name: str) -> None:
        self.en_name_item = new_en_name

    def get_en_name_item(self) -> str:
        return self.en_name_item

    def set_ua_name_item(self, new_ua_name: int) -> None:
        self.ua_name_item = new_ua_name

    def get_ua_name_item(self) -> str:
        return self.ua_name_item

    def set_amount_item(self, new_amount_item: int) -> None:
        self.amount_item = new_amount_item

    def get_amount_item(self) -> int:
        return self.amount_item

    def get_name_for_table(self) -> str:
        result: str = ""
        list_1: list = self.get_ua_name_item().split(";")
        a = list_1[0]
        b = list_1[-1]
        result = a
        result += b
        return result

    def get_sectioned_image_name(self) -> str:
        if self.get_type_holder() == "Amada-promecam":
            if self.get_type_item() == "Пуансон":
                 return "punch_amada.png"
            if self.get_type_item() == "Матриця":
                return "die_amada.png"
            if self.get_type_item() == "Пуансон плющення":
                return "punch_amada_hemming.png"
            if self.get_type_item() == "Матриця плющення":
                return "die_amada_hamming.png"
            if self.get_type_item() == "Матриця багаторучова":
             return "die_amada.png"
        if self.get_type_holder() == "Trumpf-Wila":
            if self.get_type_item() == "Пуансон":
                return "punch_trumpf.png"
            if self.get_type_item() == "Матриця":
                return "die_trumpf.png"
            if self.get_type_item() == "Пуансон плющення":
                return "punch_trumpf_hamming.png"
            if self.get_type_item() == "Матриця плющення":
                return "die_trumpf_hamming.png"
        if self.get_type_holder() == "Bystronic":
            if self.get_type_item() == "Пуансон":
                return "punch_bystronic.png"
            if self.get_type_item() == "Матриця":
                return "die_trumpf.png"
            if self.get_type_item() == "Пуансон плющення":
                return "punch_bystronic_hemming.png"
            if self.get_type_item() == "Матриця плющення":
                return "die_bystronic_hemming.png"

class Invoice:

    def __init__(
            self,
            rate: float = 0.0,
            list_item: list = None,
            packing_price: float = 0.0,
            delivery_price: float = 0.0,
            max_length: str = "0.0",
            total_weight: float = 0.0,
            commission_percentage: str = "0.0",
            provider_discount: str = "0.0",
            customer_discount: str = "0.0",
            customer_name: str = "",
            price_document: str = "",
            transaction_price: str = "",
            brokerage_price: str = "",
            bank_tax: str = "",
            total_price_ua: float = 0.0,
            total_delivery_price_ua: float = 0.0,
            sum_item_price: float = 0.0
    ) -> None:
        self.rate = rate
        self.list_item = list_item
        self.total_weight = total_weight
        self.max_length = max_length
        self.packing_price = packing_price
        self.delivery_price = delivery_price
        self.commission_percentage = commission_percentage
        self.provider_discount = provider_discount
        self.customer_discount = customer_discount
        self.customer_name = customer_name
        self.price_document = price_document
        self.transaction_price = transaction_price
        self.brokerage_price = brokerage_price
        self.bank_tax = bank_tax
        self.total_price_ua = total_price_ua
        self.sum_item_price = sum_item_price

#tecnostamp_discount

    def set_rate(self, new_rate: float) -> None:
        self.rate = new_rate

    def get_rate(self) -> float:
        return self.rate

    def set_list_item(self, new_list_item) -> None:
        self.list_item = new_list_item

    def get_list_item(self) -> list:
        if self.list_item is None:
            return []
        else:
            return self.list_item

    def add_item_to_list(self, new_item: Item) -> None:
        my_list_item = self.get_list_item()

        my_list_item.append(new_item)
        self.set_list_item(my_list_item)
        self.set_total_weight()
        self.set_max_length()

    def remove_item_from_list(self, code: str) -> None:
        index_code = 0
        for i in range(0, len(self.get_list_item())):
            if self.get_list_item()[i].get_code_item() == code:
                index_code = i
                break
        temp_list = self.get_list_item()
        temp_list.pop(index_code)
        self.set_list_item(temp_list)

    def print_code_amount(self):
        for item in self.list_item:
            print(item.get_code_item(), " ", item.get_amount_item())

    #Загальна вага
    def set_total_weight(self) -> None:
        self.total_weight = 0.0
        if self.list_item is None:
            self.total_weight = 0.0
        else:
            for item in self.list_item:
                self.total_weight += (item.get_weight_item() * item.get_amount_item())
                self.total_weight = round(self.total_weight, 2)

    def get_total_weight(self) -> float:
        return self.total_weight

    #Максимальна довжина
    def set_max_length(self) -> None:
        if len(self.list_item) == 0:
            self.max_length = "0.0"
        elif len(self.list_item) == 1:
            self.max_length = self.list_item[0].get_length_item_mm()
            self.max_length = str(float(self.max_length) / 10)
        else:
            self.max_length = self.list_item[0].get_length_item_mm()
            for i in range(len(self.list_item)):
                if float(self.list_item[i].get_length_item_mm()) > float(self.max_length):
                    self.max_length = self.list_item[i].get_length_item_mm()

            self.max_length = str(float(self.max_length) / 10)
    def get_max_length(self) -> str:
        return self.max_length

    def set_packing_price(self, new_packing_price: str) -> None:
        self.packing_price = new_packing_price

    def get_packing_price(self) -> str:
        return self.packing_price

    def set_delivery_price(self, new_delivery_price: str) -> None:
        self.delivery_price = new_delivery_price

    def get_delivery_price(self) -> str:
        return self.delivery_price

    def set_commission_percentage(self, commission: str) -> None:
        self.commission_percentage = commission

    def get_commission_percentage(self) -> str:
        return self.commission_percentage

    def set_provider_discount(self, new_provider_discount: str) -> None:
        self.provider_discount = new_provider_discount

    def get_provider_discount(self) -> str:
        return self.provider_discount

    def show_list(self) -> None:
        print("START LIST")
        for item in self.get_list_item():
            print(f"{item.get_code_item()} - {item.get_amount_item()}")
        print("END LIST")

    def get_list_code(self) -> list:
        result_list = []
        for i in self.get_list_item():
            result_list.append(i.get_code_item())
        return result_list

    def set_customer_discount(self, new_cus_discount: str) -> None:
        self.customer_discount = new_cus_discount

    def get_customer_discount(self) -> str:
        return self.customer_discount

    def set_customer_name(self, new_customer_name: str)-> None:
        self.customer_name = new_customer_name

    def get_customer_name(self) -> str:
        return self.customer_name

    def set_price_document(self, new_price_documen: str) -> None:
        self.price_document = new_price_documen

    def get_price_document(self) -> str:
        return self.price_document

    def set_transaction_price(self, new_transaction_price: str) -> None:
        self.transaction_price = new_transaction_price

    def get_transaction_price(self) -> str:
        return self.transaction_price

    def set_brokerage_price(self, new_brokerage: str) -> None:
        self.brokerage_price = new_brokerage

    def get_brokerage_price(self) -> str:
        return self.brokerage_price

    def set_bank_tax(self, new_tax: str) -> None:
        self.bank_tax = new_tax

    def get_bank_tax(self) -> str:
        return self.bank_tax

    def set_total_price_ua(self, new_price: float) -> None:
        self.total_price_ua = new_price

    def get_total_price_ua(self) -> float:
        return self.total_price_ua

    def set_total_delivery_price_ua(self, new_price: float) -> None:
        self.total_delivery_price_ua = new_price

    def get_total_delivery_price_ua(self) -> float:
        return self.total_delivery_price_ua

    def set_sum_item_price(self, price: float) -> None:
        self.sum_item_price = price

    def get_sum_item_price(self) -> float:
        return self.sum_item_price

    def calculate_sum_item_price(self) -> None:
        discount: float = (
            100 - float(self.get_provider_discount().replace(",","."))
        ) / 100
        print("Discount provider:", discount)
        sum_item_price = round(
            sum(
                [item.get_price_item() * discount * item.get_amount_item()
                 for item in self.get_list_item()]
            ),
            2)

        print("Сума закупки у виробника", sum_item_price)
        self.set_sum_item_price(sum_item_price)

    def calculate_total_price_ua(self) -> None:

        print("i", "_!"*20)
        discount: float = (
            100 - float(self.get_provider_discount().replace(",","."))
        ) / 100

        #Сума ціна * кількість * знижка постачальника
        price = round(
            sum(
                [item.get_price_item() * discount * item.get_amount_item()
                 for item in self.get_list_item()]
            ),
            2)
        print("SUM(Price * amount * discount)", price)


        #Додаємо Packing
        price += float(self.get_packing_price().replace(",", "."))
        print("Price + packing" , price)

        #Додаємо банківські відсотки
        tax: float = round((float(self.get_bank_tax().replace(",", "."))/100 ), 4)
        print(" Tax", tax)
        price = round((price * (1 + tax)), 2)
        print("Price after tax", price)

        # Додаємо комісію Vectortool
        commission: float = (
            round(
                (float(
                    self.get_commission_percentage().replace(",", "."))/100 ),
                2)
        )
        print("Commission", commission)

        price = round(price / (1 - commission), 2)
        print("Price after commission", price)

        #Переводимо у UAH
        print("rate", self.get_rate(), " type ", type(self.get_rate()) )

        price = round(price* self.get_rate(), 2)
        print("Price in UAH", price, "UAH.")

        #Додаємо Вартість переводу валюти  Брокерські
        price += float(self.get_transaction_price().replace(",", "."))
        price += float(self.get_brokerage_price().replace(",", "."))
        print("Total price UAH:",  price)

        print("i", "_!"*20)

        self.set_total_price_ua(price)

    # def calculate_total_price_ua_with_customer_discount(self) -> None:
    #     print("i", "_!"*20)
    #     discount: float = (
    #         100 - float(self.get_provider_discount().replace(",","."))
    #     ) / 100
    #
    #     #Сума ціна * кількість * знижка постачальника
    #     price = round(
    #         sum(
    #             [item.get_price_item() * discount * item.get_amount_item()
    #              for item in self.get_list_item()]
    #         ),
    #         2)
    #     print("SUM(Price * amount * discount)", price)
    #
    #
    #     #Додаємо Packing
    #     price += float(self.get_packing_price().replace(",", "."))
    #     print("Price + packing" , price)
    #
    #     #Додаємо банківські відсотки
    #     tax: float = round((float(self.get_bank_tax().replace(",", "."))/100 ), 4)
    #     print(" Tax", tax)
    #     price = round((price * (1 + tax)), 2)
    #     print("Price after tax", price)
    #
    #     # Додаємо комісію Vectortool
    #     commission: float = (
    #         round(
    #             (float(
    #                 self.get_commission_percentage().replace(",", "."))/100 ),
    #             2)
    #     )
    #     print("Commission", commission)
    #
    #     price = round(price / (1 - commission), 2)
    #     print("Price after commission", price)
    #
    #     #Знижка для клієнта
    #     customer_discount: float = round(price * (float(self.get_customer_discount().replace(",", "."))/100), 2)
    #     print("customer_discount:", customer_discount, " EURO")
    #     price = round(price - customer_discount, 2)
    #     print("Price after customer discount", price)
    #     self.set_customer_discount(str(customer_discount))
    #
    #     #Переводимо у UAH
    #     print("rate", self.get_rate(), " type ", type(self.get_rate()) )
    #
    #     price = round(price* self.get_rate(), 2)
    #     print("Price in UAH", price, "UAH.")
    #
    #     #Додаємо Вартість переводу валюти  Брокерські
    #     price += float(self.get_transaction_price().replace(",", "."))
    #     price += float(self.get_brokerage_price().replace(",", "."))
    #     print("Total price UAH:",  price)
    #
    #     print("i", "_!"*20)
    #

    #Вартість доставки у UAH
    def calculate_total_delivery_price_ua(self) -> None:

        delivery: float =\
            float(self.get_delivery_price().replace(",","."))
        document: float =\
            float(self.get_price_document().replace(",","."))
        print("delivery total", round(
            self.get_rate() * (
                    delivery + document + (delivery + document) * 0.2),
            2))

        delivery_price = round(
            self.get_rate() * (
                    delivery + document + (delivery + document) * 0.2),
            2)

        self.set_total_delivery_price_ua(delivery_price)


    def invoice_input_toString(self) -> None:
        print("#"*5, "INVOICE OBJECT","#"*5)
        print("Customer name is ", self.get_customer_name() + ".")
        print("Rate ", self.get_rate(), " uah/euro.")
        print("Commission ", self.get_commission_percentage(), "%.")
        print("Discount for customer ", self.get_customer_discount(), "%.")
        print("Provider discount", self.get_provider_discount(), "%.")
        print("Bank tax", self.get_bank_tax(), "%.")
        print("Packing price", self.get_packing_price(), "euro.")
        print("Delivery price", self.get_delivery_price(), "euro.")
        print("Document price", self.get_price_document(), "euro.")
        print("Transfer price", self.get_transaction_price(), "uah.")
        print("Brokerage price", self.get_brokerage_price(), "uah.")
        print("#"*25)

class Pre_commercial_offer:

    def __init__(
            self,
            company_name: str = None,
            rate: str = None,
            discount: str = None,
            path_temp: str = None
    ):
        self.company_name: str = company_name
        self.rate: str = rate
        self.discount: str = discount
        self.path_temp: str = path_temp

    def set_company_name(self, new_company: str) -> None:
        self.company_name = new_company

    def get_company_name(self) -> str:
        return self.company_name

    def set_rate(self, new_rate: str) -> None:
        self.rate = new_rate

    def get_rate(self) -> str:
        return self.rate

    def set_discount(self, new_discount: str) -> None:
        self.discount = new_discount

    def get_discoiunt(self) -> str:
        return self.discount

    def set_path_temp(self, new_path: str) -> None:
        self.path_temp = new_path

    def get_path_temp(self) -> str:
        return self.path_temp

    #def fill_xlsx(self, new_invoice: Invoice, new_path: str) -> None:
    def fill_xlsx(self, new_invoice: Invoice) -> None:
        #Шрифти
        #Позиція
        position_font = Font(size=6, bold=True)
        position_font.name = "Times New Roman"

        #Назва
        name_font = Font(size=7, bold=False)
        name_font.name = "Times New Roman"


        #Рамка
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        #wb = load_workbook(new_path)
        wb = load_workbook(self.get_path_temp())
        ws = wb.active
        work_sheet = wb["Лист1"]

        #Назва компанії
        #work_sheet.merge_cells(start_row=10, start_column=13, end_row=10, end_column=16)

        start_row = 17


        work_sheet.insert_rows(len(new_invoice.get_list_item()))


        current_row = 0
        last_row = len(new_invoice.get_list_item()) + start_row


        img = None
        for index in range(0, len(new_invoice.get_list_item())):
            current_row = start_row + index
            work_sheet.row_dimensions[current_row].height = 90

            #Номер позиції

            work_sheet[f"B{str(current_row)}"].font = position_font
            work_sheet[f"B{str(current_row)}"].alignment = \
                Alignment(horizontal="center", vertical='center')
            work_sheet[f"B{str(current_row)}"].value = index + 1
            work_sheet[f"B{str(current_row)}"].border = thin_border

            work_sheet[f"C{str(current_row)}"].border = thin_border

            #Англійська назва
            work_sheet[f"D{str(current_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
            work_sheet[f"D{str(current_row)}"].value = new_invoice.get_list_item()[index].get_en_name_item()
            work_sheet[f"D{str(current_row)}"].font = name_font
            work_sheet[f"D{str(current_row)}"].alignment = Alignment(horizontal="left", vertical='center',wrapText=True)
            work_sheet[f"D{str(current_row)}"].border = thin_border

            work_sheet[f"E{str(current_row)}"].border = thin_border

            #Назва українською
            work_sheet[f"F{str(current_row)}"].value = new_invoice.get_list_item()[index].get_ua_name_item()
            work_sheet[f"F{str(current_row)}"].font = name_font
            work_sheet[f"F{str(current_row)}"].alignment = Alignment(horizontal="left", vertical='center', wrapText=True)
            work_sheet[f"F{str(current_row)}"].border = thin_border

            work_sheet[f"G{str(current_row)}"].border = thin_border
            print(new_invoice.get_list_item()[index].get_image_path())
            img = openpyxl.drawing.image.Image(f"data/{new_invoice.get_list_item()[index].get_image_path()}")
            img.height = 120
            img.width = 80
            print(type(img))
            img.anchor = f"H{str(current_row)}"
            ws[f"H{str(current_row)}"].alignment = Alignment(horizontal='center', vertical="center")
            ws.add_image(img)


            #Вага
            work_sheet[f"J{str(current_row)}"].value = new_invoice.get_list_item()[index].get_weight_item()
            work_sheet[f"J{str(current_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
            work_sheet[f"J{str(current_row)}"].font = name_font
            work_sheet[f"J{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"J{str(current_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
            work_sheet[f"J{str(current_row)}"].border = thin_border

            #Кількість
            work_sheet[f"K{str(current_row)}"].value = new_invoice.get_list_item()[index].get_amount_item()
            work_sheet[f"K{str(current_row)}"].font = name_font
            work_sheet[f"K{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"K{str(current_row)}"].border = thin_border

            #Вага помножена на кількість
            work_sheet[f"I{str(current_row)}"].value = f"=J{str(current_row)}*K{str(current_row)}"
            work_sheet[f"I{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"I{str(current_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
            work_sheet[f"I{str(current_row)}"].border = thin_border

            #Закупка з урахуванням знижки
            work_sheet[f"L{str(current_row)}"].value = f"={new_invoice.get_list_item()[index].get_price_item()}*((100-{self.get_discoiunt()})/100)"
            work_sheet[f"L{str(current_row)}"].font = name_font
            work_sheet[f"L{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"L{str(current_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00', end_color='ffff00')
            work_sheet[f"L{str(current_row)}"].border = thin_border

            # Ціна за од ГРН
            work_sheet[f"O{str(current_row)}"].value = 0
            work_sheet[f"O{str(current_row)}"].font = name_font
            work_sheet[f"O{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"O{str(current_row)}"].border = thin_border

            #Ціна за од EURO
            work_sheet[f"M{str(current_row)}"].value = f"=O{str(current_row)}/{float(self.get_rate().replace(',','.'))}"
            work_sheet[f"M{str(current_row)}"].font = name_font
            work_sheet[f"M{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"M{str(current_row)}"].border = thin_border

            #Ціна разом EURO
            work_sheet[f"N{str(current_row)}"].value = f"=M{str(current_row)}*K{str(current_row)}"
            work_sheet[f"N{str(current_row)}"].font = name_font
            work_sheet[f"N{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"N{str(current_row)}"].border = thin_border

            #Ціна разом ГРН
            work_sheet[f"P{str(current_row)}"].value = f"=O{str(current_row)}*K{str(current_row)}"
            #work_sheet[f"P{str(current_row)}"].value.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
            work_sheet[f"P{str(current_row)}"].font = name_font
            work_sheet[f"P{str(current_row)}"].alignment = Alignment(horizontal="center", vertical='center')
            work_sheet[f"P{str(current_row)}"].border = thin_border

        img = None

        #last_row += 1
        work_sheet[f"B{str(last_row)}"].value = ""
        work_sheet[f"B{str(last_row)}"].border = thin_border
        work_sheet[f"C{str(last_row)}"].value = ""
        work_sheet[f"C{str(last_row)}"].border = thin_border
        work_sheet[f"D{str(last_row)}"].value = ""
        work_sheet[f"D{str(last_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00',
                                                              end_color='ffff00')
        work_sheet[f"D{str(last_row)}"].border = thin_border
        work_sheet[f"E{str(last_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00',
                                                              end_color='ffff00')
        work_sheet[f"E{str(last_row)}"].border = thin_border
        work_sheet[f"F{str(last_row)}"].value = ""
        work_sheet[f"F{str(last_row)}"].border = thin_border
        work_sheet[f"G{str(last_row)}"].value = ""
        work_sheet[f"G{str(last_row)}"].border = thin_border
        work_sheet[f"H{str(last_row)}"].value = ""
        work_sheet[f"H{str(last_row)}"].border = thin_border

        work_sheet[f"I{str(last_row)}"].font = name_font
        work_sheet[f"I{str(last_row)}"].alignment = Alignment(horizontal="center", vertical='center')


        work_sheet[f"I{str(last_row)}"] = f"=SUM(I17:I{str(last_row-1)})"
        work_sheet[f"I{str(last_row)}"].number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]
        work_sheet[f"I{str(last_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00',
                                                              end_color='ffff00')
        work_sheet[f"I{str(last_row)}"].border = thin_border
        work_sheet[f"J{str(last_row)}"].value = ""
        work_sheet[f"J{str(last_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00',
                                                              end_color='ffff00')
        work_sheet[f"J{str(last_row)}"].border = thin_border
        work_sheet[f"K{str(last_row)}"].value = ""
        work_sheet[f"K{str(last_row)}"].border = thin_border
        work_sheet[f"L{str(last_row)}"].value = ""
        work_sheet[f"L{str(last_row)}"].fill = PatternFill(fill_type='solid', start_color='ffff00',
                                                              end_color='ffff00')
        work_sheet[f"L{str(last_row)}"].border = thin_border
        work_sheet[f"M{str(last_row)}"].value = ""
        work_sheet[f"M{str(last_row)}"].border = thin_border
        work_sheet[f"N{str(last_row)}"].value = ""
        work_sheet[f"N{str(last_row)}"].border = thin_border
        work_sheet[f"O{str(last_row)}"].value = ""
        work_sheet[f"O{str(last_row)}"].border = thin_border
        work_sheet[f"P{str(last_row)}"].value = ""
        work_sheet[f"P{str(last_row)}"].border = thin_border

        work_sheet.merge_cells('M10:P10')
        work_sheet.cell(row=10, column=13).value = get_full_name_company(self.get_company_name())
        name_company_font = Font(size=9, bold=True)
        name_company_font.name = "Times New Roman"
        work_sheet.cell(row=10, column=13).font = name_company_font
        work_sheet.cell(row=10, column=13).alignment = Alignment(horizontal="right", vertical='center')

        #Разом
        work_sheet.merge_cells(f'F{str(last_row+1)}:M{last_row+1}')
        work_sheet[f"F{last_row+1}"].value = "Разом"
        #work_sheet.cell(row=6, column=last_row+1).value = "Разом"
        all_font = Font(size=9, bold=False)
        all_font.name = "Times New Roman"


        work_sheet[f"F{last_row + 1}"].font = all_font
        work_sheet[f"F{last_row + 1}"].alignment = Alignment(horizontal="left", vertical='center')
        print(str(last_row+1))


        #Загальна ціна у EURO та грн
        if len(new_invoice.get_list_item()) == 1:
            work_sheet[f"N{str(last_row + 1)}"] = f"=N{str(start_row)}"

            work_sheet[f"P{str(last_row + 1)}"] = f"=P{str(start_row)}"
        else:
           # work_sheet[f"N{str(last_row + 1)}"] = f'=СУММ(N{str(start_row)}:N{str(last_row)})'
            work_sheet[f"N{str(last_row + 1)}"] = f'=SUM(N{str(start_row)}:N{str(last_row)})'
            work_sheet[f"P{str(last_row + 1)}"] = f"=SUM(P{str(start_row)}:P{str(last_row)})"
        work_sheet[f"F{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"G{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"H{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"I{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"J{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"K{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"L{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"M{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"N{str(last_row + 1)}"] .border = thin_border
        work_sheet[f"P{str(last_row + 1)}"] .border = thin_border

        #ПДВ
        work_sheet.merge_cells(f'F{str(last_row+2)}:M{last_row+2}')
        work_sheet[f"F{last_row+2}"].value = "Податок на додану вартість (ПДВ)"
        work_sheet[f"F{last_row + 2}"].font = all_font
        work_sheet[f"F{last_row + 2}"].alignment = Alignment(horizontal="left", vertical='center')
        work_sheet[f"F{last_row + 2}"].border = thin_border

        work_sheet[f"G{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"H{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"I{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"J{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"K{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"L{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"M{str(last_row + 2)}"] .border = thin_border

        work_sheet[f"N{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"N{str(last_row + 2)}"] = f"=N{str(last_row + 1)}*0.2"

        work_sheet[f"P{str(last_row + 2)}"] .border = thin_border
        work_sheet[f"P{str(last_row + 2)}"] = f"=P{str(last_row + 1)}*0.2"

        #Загальна вартість з ПДВ
        work_sheet.merge_cells(f'F{str(last_row+3)}:M{last_row+3}')
        work_sheet[f"F{last_row+3}"].value = "Загальна вартість з ПДВ"
        work_sheet[f"F{last_row + 3}"].font = all_font
        work_sheet[f"F{last_row + 3}"].alignment = Alignment(horizontal="left", vertical='center')
        work_sheet[f"F{last_row + 3}"].border = thin_border

        work_sheet[f"G{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"H{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"I{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"J{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"K{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"L{str(last_row + 3)}"] .border = thin_border

        work_sheet[f"M{str(last_row + 3)}"] .border = thin_border
        work_sheet[f"N{str(last_row + 3)}"] = f"=N{str(last_row + 1)}+N{str(last_row + 2)} "
        work_sheet[f"N{str(last_row + 3)}"] .border = thin_border

        work_sheet[f"P{str(last_row + 3)}"] = f"=P{str(last_row + 1)}+P{str(last_row + 2)} "
        work_sheet[f"P{str(last_row + 3)}"] .border = thin_border


        #Рамка порожніх ячеек
        work_sheet[f"O{last_row + 1}"].border = thin_border
        work_sheet[f"O{last_row + 2}"].border = thin_border
        work_sheet[f"O{last_row + 3}"].border = thin_border

        current_row = last_row + 3
        if new_invoice.get_customer_discount() != 0:
            work_sheet.merge_cells(
                f'F{str(current_row)}:M{current_row}'
            )
            work_sheet[f"F{current_row}"].value = \
                f"Знижка для компанії "


        wb.save(self.get_path_temp())


